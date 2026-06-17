import os
import shutil
import sys
from datetime import datetime

from fpdf import FPDF
from PyQt6.QtCore import QDate, QSize, Qt
from PyQt6.QtGui import QFont, QIcon, QPixmap
from PyQt6.QtWidgets import (
    QApplication,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from app_logger import AppLogger
from constants import PAGE_SIZE_DEFAULT, PASS_AVERAGE, STUDENT_CODE_PREFIX
from database_setup import DatabaseManager
from print_export_service import get_report_output_mode, output_pdf
from repositories.finance_repo import FinanceRepository
from repositories.student_repo import StudentRepository
from ui_components import (
    card_frame,
    compact_icon_btn,
    horizontal_separator,
    section_label,
    style_table,
    styled_button,
    styled_combo,
    styled_date_edit,
    styled_input,
    vertical_separator,
)
from ui_styles import (
    EmptyStateWidget,
    ModuleHeaderWidget,
    PaginationWidget,
    ThemeManager,
    ToastNotification,
    friendly_db_error,
    get_module_caps,
    get_tabs_style,
)

STUDENT_LIST_OUTPUT_MODE = get_report_output_mode("student_list_mode", "save")

from pdf_helpers import (
    ARABIC_SUPPORT,
    is_arabic_font_ready,
    latin_fallback,
    prepare_pdf_text,
    sanitize_latin,
    setup_pdf_arabic_font,
)

# --- فئة توليد PDF (كما هي) ---


class StudentListPDF(FPDF):
    def __init__(self, school_info=None, title_doc="LISTE DES ETUDIANTS", orientation='P'):
        super().__init__(orientation=orientation, unit='mm', format='A4')
        self.school_info = school_info
        self.title_doc = title_doc

    def sanitize(self, text):
        return sanitize_latin(text)

    def header(self):
        left_x, left_y = 10, 5
        page_w = self.w
        right_x = page_w - 30
        self.set_xy(left_x, left_y)
        self.set_font('Helvetica', '', 8)
        self.set_text_color(30, 41, 59)

        if self.school_info and len(self.school_info) > 7:
            republic = self.sanitize(self.school_info[1])
            self.cell(80, 3, republic, 0, 1, 'L')
            ia_text = self.sanitize(self.school_info[2])
            self.cell(80, 3, ia_text, 0, 1, 'L')
            ief_text = self.sanitize(self.school_info[3])
            self.cell(80, 3, ief_text, 0, 1, 'L')
            school_name = self.sanitize(self.school_info[4])
            self.cell(80, 3, school_name, 0, 1, 'L')
            auth_text = self.sanitize(self.school_info[5])
            self.cell(80, 3, f"Auto N: {auth_text}", 0, 1, 'L')
            addr_text = self.sanitize(self.school_info[6])
            self.cell(80, 3, f"Lieu: {addr_text}", 0, 1, 'L')
            phone_text = self.sanitize(self.school_info[7])
            self.cell(80, 3, f"Tel: {phone_text}", 0, 1, 'L')

        logo_path = self.school_info[8] if self.school_info and len(self.school_info) > 8 else None
        if logo_path and os.path.exists(logo_path):
            try:
                self.image(logo_path, x=right_x, y=left_y, w=20, h=22)
            except Exception:
                pass

        # Keep divider/title position stable even when school info content varies.
        body_bottom_y = self.get_y()
        line_y = max(body_bottom_y + 2, left_y + 24)
        self.line(10, line_y, page_w - 10, line_y)
        self.set_y(line_y + 4)

        self.set_font('Helvetica', 'B', 12)
        self.set_fill_color(30, 41, 59)  # Slate 800
        self.set_text_color(248, 250, 252)  # Slate 50
        self.cell(0, 8, self.title_doc, 0, 1, 'C', True)
        self.set_text_color(44, 62, 80)
        self.ln(2)

    def footer(self):
        self.set_y(-12)
        self.set_font('Helvetica', 'I', 7)
        date_str = datetime.now().strftime('%Y-%m-%d')
        page_w = self.w - self.l_margin - self.r_margin
        self.cell(page_w / 2, 4, f"Imprimé le {date_str}", 0, 0, 'L')
        self.cell(page_w / 2, 4, f"Page {self.page_no()}", 0, 0, 'R')


# ─────────────────────────────────────────────────────────────────────────────
#  StudentDialog — نافذة إضافة/تعديل الطالب المنبثقة
# ─────────────────────────────────────────────────────────────────────────────


class StudentDialog(QDialog):
    """QDialog لإضافة أو تعديل بيانات طالب.

    المعاملات:
        cycles_data : list[tuple(id, name)]  — دورات التعليم المحملة مسبقاً
        get_classes_fn : callable(cycle_id) -> list[tuple(id, name)]
        get_next_number_fn : callable(class_id) -> int
        data : dict | None  — بيانات الطالب الحالية عند التعديل
        parent : QWidget | None
    """

    # كود الخروج لزر «حفظ وجديد»
    RESULT_SAVE_AND_NEW = 10

    def __init__(self, cycles_data, get_classes_fn, get_next_number_fn, data=None, parent=None):
        super().__init__(parent)
        self.cycles_data = cycles_data
        self._get_classes_fn = get_classes_fn
        self._get_next_number_fn = get_next_number_fn
        self.data = data  # dict أو None
        self.current_photo_path = data.get("photo_path") if data else None
        self._is_edit = data is not None

        self.setWindowTitle("✏️  Modifier l'élève" if self._is_edit else "➕  Nouvel Élève / طالب جديد")
        self.setModal(True)
        self.setMinimumWidth(720)
        self.setMinimumHeight(560)
        ThemeManager.apply_theme(self)

        self._build_ui()
        if self._is_edit:
            self._populate(data)
        self._refresh_tab_indicators()

    # ── بناء الواجهة ────────────────────────────────────────────────────────

    def _build_ui(self):
        colors = ThemeManager.get_colors()
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(10)

        # ── تبويبات ──────────────────────────────────────────────────────────
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(get_tabs_style())
        self._build_tab_identity(colors)
        self._build_tab_placement(colors)
        self._build_tab_parent(colors)
        self.tabs.currentChanged.connect(self._on_tab_changed)
        root.addWidget(self.tabs)

        # ── صف التنقل ────────────────────────────────────────────────────────
        nav_row = QHBoxLayout()
        nav_row.setSpacing(8)

        self._btn_prev = styled_button(
            "← Précédent",
            bg_color="transparent",
            text_color=colors.TEXT_SECONDARY,
            hover_color=colors.BG_MAIN,
            min_height=36,
        )
        self._btn_prev.setEnabled(False)
        self._btn_prev.setStyleSheet(
            self._btn_prev.styleSheet()
            + f"QPushButton:disabled {{ color:{colors.BORDER}; border:1.5px solid {colors.BORDER}; }}"
        )
        self._btn_prev.clicked.connect(self._go_prev)

        self._btn_next = styled_button("Suivant →", min_height=36)
        self._btn_next.clicked.connect(self._go_next)

        # فاصل مرئي
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setStyleSheet(f"color:{colors.BORDER};")

        # زر «حفظ وجديد» — في وضع الإضافة فقط
        self._btn_save_new = None
        if not self._is_edit:
            self._btn_save_new = QPushButton("💾+  Enregistrer & Nouveau")
            self._btn_save_new.setMinimumHeight(36)
            self._btn_save_new.setCursor(Qt.CursorShape.PointingHandCursor)
            self._btn_save_new.setStyleSheet(
                f"QPushButton {{ background:transparent; color:{colors.PRIMARY};"
                f" font-weight:700; font-size:12px; border-radius:7px;"
                f" border:2px solid {colors.PRIMARY}; padding:6px 16px; }}"
                f"QPushButton:hover {{ background:{colors.PRIMARY_LIGHT}; }}"
            )
            self._btn_save_new.clicked.connect(self._on_accept_and_new)

        save_label = "✏️  Modifier" if self._is_edit else "💾  Enregistrer"
        self._btn_save = styled_button(
            save_label, bg_color=colors.SUCCESS, hover_color=colors.SUCCESS_HOVER, min_height=36
        )
        self._btn_save.clicked.connect(self._on_accept)

        btn_cancel = styled_button(
            "✕  Annuler",
            bg_color="transparent",
            text_color=colors.TEXT_SECONDARY,
            hover_color=colors.BG_MAIN,
            min_height=36,
        )
        btn_cancel.clicked.connect(self.reject)

        nav_row.addWidget(self._btn_prev)
        nav_row.addWidget(self._btn_next)
        nav_row.addWidget(sep)
        nav_row.addStretch()
        if self._btn_save_new:
            nav_row.addWidget(self._btn_save_new)
        nav_row.addWidget(self._btn_save)
        nav_row.addWidget(btn_cancel)
        root.addLayout(nav_row)

    def _make_input(self, placeholder="", read_only=False):
        return styled_input(placeholder, min_height=38, read_only=read_only)

    def _make_combo(self):
        return styled_combo(min_height=38)

    def _make_date(self):
        return styled_date_edit(min_height=38)

    def _section_label(self, icon, text):
        return section_label(icon, text)

    # ── التبويب 1: البيانات الشخصية ─────────────────────────────────────────

    def _build_tab_identity(self, colors):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        # صورة + حقول الاسم في صف واحد
        top_row = QHBoxLayout()
        top_row.setSpacing(16)

        # صورة
        photo_col = QVBoxLayout()
        photo_col.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)
        self.lbl_photo = QLabel()
        self.lbl_photo.setFixedSize(90, 90)
        self.lbl_photo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_photo.setText("📷")
        self.lbl_photo.setStyleSheet(
            f"QLabel {{ background:{colors.PRIMARY_LIGHT}; border-radius:45px;"
            f"border:3px solid {colors.PRIMARY}; color:{colors.TEXT_SECONDARY};"
            f"font-size:24px; }}"
        )
        btn_photo = QPushButton("Changer")
        btn_photo.setMinimumHeight(28)
        btn_photo.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_photo.setStyleSheet(
            f"QPushButton {{ background:transparent; color:{colors.PRIMARY};"
            f"font-size:11px; font-weight:600; border:1px solid {colors.PRIMARY};"
            f"border-radius:5px; padding:3px 8px; }}"
            f"QPushButton:hover {{ background:{colors.PRIMARY_LIGHT}; }}"
        )
        btn_photo.clicked.connect(self._upload_photo)
        photo_col.addWidget(self.lbl_photo, 0, Qt.AlignmentFlag.AlignHCenter)
        photo_col.addWidget(btn_photo, 0, Qt.AlignmentFlag.AlignHCenter)
        top_row.addLayout(photo_col)

        # حقول الاسم (2×2 grid)
        name_grid = QGridLayout()
        name_grid.setSpacing(8)
        name_grid.addWidget(QLabel("Prénom (FR):"), 0, 0)
        self.txt_fname_fr = self._make_input("Prénom")
        name_grid.addWidget(self.txt_fname_fr, 0, 1)
        name_grid.addWidget(QLabel("Nom (FR):"), 1, 0)
        self.txt_lname_fr = self._make_input("Nom de famille")
        name_grid.addWidget(self.txt_lname_fr, 1, 1)
        name_grid.addWidget(QLabel("الاسم (AR):"), 2, 0)
        self.txt_fname_ar = self._make_input("الاسم")
        name_grid.addWidget(self.txt_fname_ar, 2, 1)
        name_grid.addWidget(QLabel("اللقب (AR):"), 3, 0)
        self.txt_lname_ar = self._make_input("اللقب")
        name_grid.addWidget(self.txt_lname_ar, 3, 1)
        top_row.addLayout(name_grid, 1)
        layout.addLayout(top_row)

        # بيانات إضافية
        layout.addWidget(self._section_label("📋", "Informations complémentaires"))
        grid2 = QGridLayout()
        grid2.setSpacing(8)

        grid2.addWidget(QLabel("Naissance:"), 0, 0)
        self.date_birth = self._make_date()
        self.date_birth.setDate(QDate(2015, 1, 1))
        grid2.addWidget(self.date_birth, 0, 1)

        grid2.addWidget(QLabel("Lieu:"), 0, 2)
        self.txt_birth_place = self._make_input("Lieu de naissance")
        grid2.addWidget(self.txt_birth_place, 0, 3)

        grid2.addWidget(QLabel("Sexe:"), 1, 0)
        self.combo_gender = self._make_combo()
        self.combo_gender.addItems(["Masculin", "Féminin"])
        grid2.addWidget(self.combo_gender, 1, 1)

        grid2.addWidget(QLabel("Adresse:"), 1, 2)
        self.txt_address = self._make_input("Adresse complète")
        grid2.addWidget(self.txt_address, 1, 3)

        layout.addLayout(grid2)
        # ربط مؤشرات الإكمال بتغيّر حقلَي الاسم الإلزاميَّين
        self.txt_fname_fr.textChanged.connect(self._refresh_tab_indicators)
        self.txt_lname_fr.textChanged.connect(self._refresh_tab_indicators)
        layout.addStretch()
        self.tabs.addTab(tab, "👤  Identité")

    # ── التبويب 2: التنسيب ──────────────────────────────────────────────────

    def _build_tab_placement(self, colors):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        layout.addWidget(self._section_label("🎓", "Affectation / التنسيب"))

        grid = QGridLayout()
        grid.setSpacing(8)

        grid.addWidget(QLabel("Cycle:"), 0, 0)
        self.cmb_cycle = self._make_combo()
        self.cmb_cycle.addItem("Choisir cycle...", None)
        for cid, cname in self.cycles_data:
            self.cmb_cycle.addItem(cname, cid)
        self.cmb_cycle.currentIndexChanged.connect(self._on_cycle_changed)
        grid.addWidget(self.cmb_cycle, 0, 1)

        grid.addWidget(QLabel("Classe:"), 1, 0)
        self.cmb_class = self._make_combo()
        self.cmb_class.addItem("Choisir classe...", None)
        self.cmb_class.currentIndexChanged.connect(self._on_class_changed)
        self.cmb_class.currentIndexChanged.connect(self._refresh_tab_indicators)
        grid.addWidget(self.cmb_class, 1, 1)

        grid.addWidget(QLabel("N° Classe:"), 2, 0)
        self.txt_class_number = self._make_input("Auto")
        self.txt_class_number.setReadOnly(True)
        self.txt_class_number.setStyleSheet(
            f"QLineEdit {{ padding:7px 12px; border:1.5px solid {colors.BORDER_FOCUS};"
            f"border-radius:8px; background:{colors.PRIMARY_LIGHT}; color:{colors.PRIMARY};"
            f"font-weight:700; font-size:13px; }}"
        )
        grid.addWidget(self.txt_class_number, 2, 1)

        layout.addWidget(self._section_label("📋", "Statut / الحالة"))
        grid2 = QGridLayout()
        grid2.setSpacing(8)
        grid2.addWidget(QLabel("Date inscr.:"), 0, 0)
        self.date_registration = self._make_date()
        self.date_registration.setDate(QDate.currentDate())
        grid2.addWidget(self.date_registration, 0, 1)
        grid2.addWidget(QLabel("Statut:"), 1, 0)
        self.cmb_status = self._make_combo()
        self.cmb_status.addItems(["Active", "Inactive", "Suspendu", "Diplômé"])
        grid2.addWidget(self.cmb_status, 1, 1)

        layout.addLayout(grid)
        layout.addLayout(grid2)
        layout.addStretch()
        self.tabs.addTab(tab, "🎓  Affectation")

    # ── التبويب 3: معلومات الولي ─────────────────────────────────────────────

    def _build_tab_parent(self, colors):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        layout.addWidget(self._section_label("👨‍👩‍👧", "Informations du tuteur / معلومات الولي"))
        grid = QGridLayout()
        grid.setSpacing(8)

        grid.addWidget(QLabel("Nom:"), 0, 0)
        self.txt_parent_name = self._make_input("Nom du tuteur")
        grid.addWidget(self.txt_parent_name, 0, 1)

        grid.addWidget(QLabel("Tél:"), 0, 2)
        self.txt_parent_phone = self._make_input("Téléphone")
        grid.addWidget(self.txt_parent_phone, 0, 3)

        grid.addWidget(QLabel("Email:"), 1, 0)
        self.txt_parent_email = self._make_input("Email tuteur")
        grid.addWidget(self.txt_parent_email, 1, 1)

        grid.addWidget(QLabel("Adresse:"), 1, 2)
        self.txt_parent_addr = self._make_input("Adresse du tuteur")
        grid.addWidget(self.txt_parent_addr, 1, 3)

        layout.addLayout(grid)
        layout.addStretch()
        self.tabs.addTab(tab, "👨‍👩‍👧  Tuteur")

    # ── استجابات الأحداث ─────────────────────────────────────────────────────

    # ── أزرار التنقل ─────────────────────────────────────────────────────────

    def _go_prev(self):
        self.tabs.setCurrentIndex(self.tabs.currentIndex() - 1)

    def _go_next(self):
        self.tabs.setCurrentIndex(self.tabs.currentIndex() + 1)

    def _on_tab_changed(self, idx: int):
        count = self.tabs.count()
        self._btn_prev.setEnabled(idx > 0)
        self._btn_next.setVisible(idx < count - 1)
        if self._btn_save_new:
            self._btn_save_new.setVisible(idx == count - 1)
        self._btn_save.setVisible(idx == count - 1)

    def _refresh_tab_indicators(self):
        """يحدّث نصوص التبويبات ليعكس حالة الإكمال."""
        has_id = bool(self.txt_fname_fr.text().strip() and self.txt_lname_fr.text().strip())
        has_class = bool(self.cmb_class.currentData())
        self.tabs.setTabText(0, ("✓ " if has_id else "○ ") + "Identité")
        self.tabs.setTabText(1, ("✓ " if has_class else "○ ") + "Affectation")
        self.tabs.setTabText(2, "✓  Tuteur")

    # ── الحفظ وإعادة الفتح ──────────────────────────────────────────────────

    def _validate(self) -> bool:
        """يُعيد True إذا اجتاز النموذج التحقق."""
        fn_fr = self.txt_fname_fr.text().strip()
        ln_fr = self.txt_lname_fr.text().strip()
        if not fn_fr or not ln_fr:
            QMessageBox.warning(self, "Champs obligatoires", "Prénom et Nom (FR) sont obligatoires.")
            self.tabs.setCurrentIndex(0)
            self.txt_fname_fr.setFocus()
            return False
        if not self.cmb_class.currentData():
            QMessageBox.warning(self, "Classe manquante", "Veuillez sélectionner une classe.")
            self.tabs.setCurrentIndex(1)
            return False
        return True

    def _on_accept_and_new(self):
        if self._validate():
            self.done(self.RESULT_SAVE_AND_NEW)

    def _on_cycle_changed(self):
        cycle_id = self.cmb_cycle.currentData()
        self.cmb_class.blockSignals(True)
        self.cmb_class.clear()
        self.cmb_class.addItem("Choisir classe...", None)
        if cycle_id:
            for cid, cname in self._get_classes_fn(cycle_id):
                self.cmb_class.addItem(cname, cid)
        self.cmb_class.blockSignals(False)
        self._on_class_changed()

    def _on_class_changed(self):
        class_id = self.cmb_class.currentData()
        if not class_id:
            self.txt_class_number.clear()
            return
        # عند التعديل: لا نغيّر الرقم إذا كانت نفس الفصل
        if self._is_edit and self.data and self.data.get("class_id") == class_id:
            self.txt_class_number.setText(str(self.data.get("class_number", "")))
            return
        next_num = self._get_next_number_fn(class_id)
        self.txt_class_number.setText(str(next_num) if next_num else "1")

    def _upload_photo(self):
        path, _ = QFileDialog.getOpenFileName(self, "Choisir une photo", "", "Images (*.png *.jpg *.jpeg)")
        if path:
            self.current_photo_path = path
            pix = QPixmap(path).scaled(
                90, 90, Qt.AspectRatioMode.KeepAspectRatioByExpanding, Qt.TransformationMode.SmoothTransformation
            )
            self.lbl_photo.setPixmap(pix)
            self.lbl_photo.setText("")

    # ── ملء البيانات عند التعديل ─────────────────────────────────────────────

    def _populate(self, d: dict):
        self.txt_fname_fr.setText(d.get("first_name_fr") or "")
        self.txt_lname_fr.setText(d.get("last_name_fr") or "")
        self.txt_fname_ar.setText(d.get("first_name_ar") or "")
        self.txt_lname_ar.setText(d.get("last_name_ar") or "")
        if d.get("birth_date"):
            self.date_birth.setDate(QDate.fromString(str(d["birth_date"]), "yyyy-MM-dd"))
        self.txt_birth_place.setText(d.get("birth_place") or "")
        gender_val = d.get("gender", "M")
        self.combo_gender.setCurrentIndex(0 if gender_val in (None, "", "M", "m", "Masculin") else 1)
        self.txt_address.setText(d.get("address") or "")

        # تنسيب: cycle → class
        class_id = d.get("class_id")
        if class_id:
            # نبحث عن الـ cycle المناسب
            for i in range(1, self.cmb_cycle.count()):
                cid = self.cmb_cycle.itemData(i)
                classes = self._get_classes_fn(cid)
                if any(cl[0] == class_id for cl in classes):
                    self.cmb_cycle.blockSignals(True)
                    self.cmb_cycle.setCurrentIndex(i)
                    self.cmb_cycle.blockSignals(False)
                    # نملأ قائمة الفصول يدوياً
                    self.cmb_class.blockSignals(True)
                    self.cmb_class.clear()
                    self.cmb_class.addItem("Choisir classe...", None)
                    for cl_id, cl_name in classes:
                        self.cmb_class.addItem(cl_name, cl_id)
                    idx = self.cmb_class.findData(class_id)
                    if idx >= 0:
                        self.cmb_class.setCurrentIndex(idx)
                    self.cmb_class.blockSignals(False)
                    break
        self.txt_class_number.setText(str(d.get("class_number", "")))

        if d.get("registration_date"):
            self.date_registration.setDate(QDate.fromString(str(d["registration_date"]), "yyyy-MM-dd"))
        self.cmb_status.setCurrentText(d.get("status") or "Active")

        self.txt_parent_name.setText(d.get("parent_name") or "")
        self.txt_parent_phone.setText(d.get("parent_phone") or "")
        self.txt_parent_email.setText(d.get("parent_email") or "")
        self.txt_parent_addr.setText(d.get("parent_address") or "")

        photo_path = d.get("photo_path")
        if photo_path and os.path.exists(photo_path):
            self.current_photo_path = photo_path
            pix = QPixmap(photo_path).scaled(
                90, 90, Qt.AspectRatioMode.KeepAspectRatioByExpanding, Qt.TransformationMode.SmoothTransformation
            )
            self.lbl_photo.setPixmap(pix)
            self.lbl_photo.setText("")

    # ── التحقق والقبول ───────────────────────────────────────────────────────

    def _on_accept(self):
        if self._validate():
            self.accept()

    def get_values(self) -> dict:
        """يُعيد جميع قيم النموذج كـ dict."""
        return {
            "first_name_fr": self.txt_fname_fr.text().strip(),
            "last_name_fr": self.txt_lname_fr.text().strip(),
            "first_name_ar": self.txt_fname_ar.text().strip(),
            "last_name_ar": self.txt_lname_ar.text().strip(),
            "birth_date": self.date_birth.date().toString("yyyy-MM-dd"),
            "birth_place": self.txt_birth_place.text().strip(),
            "gender": "M" if self.combo_gender.currentIndex() == 0 else "F",
            "address": self.txt_address.text().strip(),
            "class_id": self.cmb_class.currentData(),
            "registration_date": self.date_registration.date().toString("yyyy-MM-dd"),
            "status": self.cmb_status.currentText(),
            "parent_name": self.txt_parent_name.text().strip(),
            "parent_phone": self.txt_parent_phone.text().strip(),
            "parent_email": self.txt_parent_email.text().strip(),
            "parent_address": self.txt_parent_addr.text().strip(),
            "photo_path": self.current_photo_path,
        }


# ─────────────────────────────────────────────────────────────────────────────


class ModernStudentManagement(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gestion des Élèves / إدارة الطلاب")
        self.setMinimumSize(1100, 700)

        # تطبيق المظهر (Dark Mode أو Light Mode)
        ThemeManager.apply_theme(self)
        self.current_photo_path = None
        self.selected_student_id = None
        # علامات RBAC — يتم تحديثهما بواسطة apply_rbac() بعد الإنشاء
        self._rbac_can_write = True
        self._rbac_can_delete = True
        self._initialized = False  # منع التحميل المزدوج عند استدعاء apply_rbac()
        self.init_ui()
        self.load_cycles_filter()
        self._load_classes_into(self.combo_filter_class_reg, None, "Toutes les Classes")
        self.refresh_student_list()
        self._load_kpi_stats()
        self._initialized = True

    def apply_rbac(self, role: str) -> None:
        """تطبيق صلاحيات الأزرار بناءً على دور المستخدم الحالي.
        يُستدعى من MainWindow بعد إنشاء الوحدة مباشرةً.
        """
        caps = get_module_caps(role, "student_management")
        self._rbac_can_write = caps["can_write"]
        self._rbac_can_delete = caps["can_delete"]
        if hasattr(self, "btn_add"):
            self.btn_add.setEnabled(caps["can_write"])
            self.btn_add.setVisible(caps["can_write"])
        if self._initialized:
            self.refresh_student_list()

    def init_ui(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(24, 20, 24, 20)
        self.main_layout.setSpacing(16)

        # 1. En-tête unifié
        header = ModuleHeaderWidget(
            icon="👨‍🎓",
            title="GESTION DES ÉLÈVES",
            subtitle="إدارة ملفات الطلاب والتسجيل",
        )
        self.main_layout.addWidget(header)
        self._stat_total = header.add_stat("👥", "Total Élèves", "—", "#3B82F6")
        self._stat_actifs = header.add_stat("✅", "Actifs", "—", "#22C55E")
        self._stat_garcons = header.add_stat("👦", "Garçons", "—", "#8B5CF6")
        self._stat_filles = header.add_stat("👧", "Filles", "—", "#EC4899")

        # 2. Statistiques rapides (KPI)

        # 3. Onglets
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(get_tabs_style())

        self.setup_student_tab()
        self.setup_list_tab()

        self.main_layout.addWidget(self.tabs)

    def setup_student_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 16, 20, 16)
        colors = ThemeManager.get_colors()

        # ── Toolbar ──────────────────────────────────────────────────────────
        toolbar_frame = card_frame()
        tlay = QHBoxLayout(toolbar_frame)
        tlay.setContentsMargins(10, 6, 10, 6)
        tlay.setSpacing(8)

        self.btn_add = QPushButton("➕  Ajouter Élève")
        self.btn_add.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_add.setFixedHeight(32)
        self.btn_add.setStyleSheet(
            f"QPushButton {{ background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            f"stop:0 {colors.SUCCESS}, stop:1 #16A34A); color:white; font-weight:700;"
            f"font-size:12px; border-radius:7px; border:none; padding:4px 14px; }}"
            f"QPushButton:hover {{ background:{colors.SUCCESS_HOVER}; }}"
            f"QPushButton:disabled {{ background:{colors.BORDER}; color:{colors.TEXT_SECONDARY}; }}"
        )
        self.btn_add.clicked.connect(lambda: self.open_student_dialog())

        self.txt_search_reg = styled_input("🔍 Recherche...", min_height=32)
        self.txt_search_reg.setMaximumWidth(250)
        self.txt_search_reg.textChanged.connect(self.refresh_student_list)

        self.combo_filter_class_reg = styled_combo(min_height=32)
        self.combo_filter_class_reg.addItem("Toutes les Classes", None)
        self.combo_filter_class_reg.setMaximumWidth(190)
        self.combo_filter_class_reg.currentIndexChanged.connect(self.refresh_student_list)

        # Toggle langue du nom: QComboBox avec icône claire
        self.combo_name_lang_reg = styled_combo(min_height=32)
        self.combo_name_lang_reg.addItem("🌐 FR", "fr")
        self.combo_name_lang_reg.addItem("🌐 AR", "ar")
        self.combo_name_lang_reg.setMaximumWidth(120)
        self.combo_name_lang_reg.setToolTip("Langue d'affichage du nom")
        self.combo_name_lang_reg.currentIndexChanged.connect(self.refresh_student_list)

        # فاصل مرئي بين الفلاتر وأزرار الإجراءات
        sep_toolbar = QFrame()
        sep_toolbar.setFrameShape(QFrame.Shape.VLine)
        sep_toolbar.setFixedHeight(22)
        sep_toolbar.setStyleSheet(f"color:{colors.BORDER};")

        btn_print_filtered = QPushButton("🖨 Imprimer")
        btn_print_filtered.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_print_filtered.setFixedHeight(32)
        btn_print_filtered.setStyleSheet(
            f"QPushButton {{ background:transparent; color:{colors.TEXT_PRIMARY};"
            f" font-weight:600; font-size:11px; border:1.5px solid {colors.BORDER};"
            f" border-radius:7px; padding:4px 10px; }}"
            f"QPushButton:hover {{ background:{colors.BG_MAIN}; border-color:{colors.TEXT_SECONDARY}; }}"
        )
        btn_print_filtered.clicked.connect(self.print_filtered_list)

        btn_import_excel = QPushButton("📥 CSV")
        btn_import_excel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_import_excel.setFixedHeight(32)
        btn_import_excel.setStyleSheet(
            f"QPushButton {{ background:transparent; color:{colors.PRIMARY};"
            f" font-weight:600; font-size:11px; border:1.5px solid {colors.PRIMARY};"
            f" border-radius:7px; padding:4px 10px; }}"
            f"QPushButton:hover {{ background:{colors.PRIMARY_LIGHT}; }}"
        )
        btn_import_excel.clicked.connect(self._open_import_wizard)

        tlay.addWidget(self.btn_add)
        tlay.addWidget(self.txt_search_reg)
        tlay.addWidget(self.combo_filter_class_reg)
        tlay.addWidget(self.combo_name_lang_reg)
        tlay.addWidget(sep_toolbar)
        tlay.addStretch()
        tlay.addWidget(btn_print_filtered)
        tlay.addWidget(btn_import_excel)
        layout.addWidget(toolbar_frame)

        # ── جدول كامل العرض ──────────────────────────────────────────────────
        self.table_students_reg = QTableWidget()
        style_table(self.table_students_reg)
        self.table_students_reg.setColumnCount(5)
        self.table_students_reg.setHorizontalHeaderLabels(["ID", "N° Classe", "Nom & Prénom", "Classe", "⚙️ Actions"])
        hh = self.table_students_reg.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.table_students_reg.setColumnWidth(4, 130)
        layout.addWidget(self.table_students_reg)

        self.tabs.addTab(tab, "  Inscription & Gestion / التسجيل والإدارة  ")

    def setup_list_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        filter_frame = card_frame()
        flay = QHBoxLayout(filter_frame)
        flay.setContentsMargins(15, 15, 15, 15)
        flay.setSpacing(15)

        self.combo_filter_cycle = styled_combo()
        self.combo_filter_cycle.addItem("Cycles (Tous)", None)
        self.combo_filter_cycle.currentIndexChanged.connect(self.load_classes_for_filter)
        self.combo_filter_cycle.setFixedWidth(120)

        self.combo_filter_class = styled_combo()
        self.combo_filter_class.addItem("Classes (Toutes)", None)
        self.combo_filter_class.currentIndexChanged.connect(self.refresh_student_list)
        self.combo_filter_class.setFixedWidth(130)

        self.date_filter_from = styled_date_edit()
        self.date_filter_from.setDate(QDate(2025, 10, 1))
        self.date_filter_from.setFixedWidth(110)
        self.date_filter_from.dateChanged.connect(self.refresh_student_list)

        self.date_filter_to = styled_date_edit()
        self.date_filter_to.setDate(QDate.currentDate())
        self.date_filter_to.setFixedWidth(110)
        self.date_filter_to.dateChanged.connect(self.refresh_student_list)

        self.txt_search = styled_input("🔍 Recherche globale...")
        self.txt_search.textChanged.connect(self.refresh_student_list)

        colors = ThemeManager.get_colors()
        btn_print = QPushButton("🖨  Imprimer")
        btn_print.setMinimumHeight(42)
        btn_print.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_print.setStyleSheet(
            f"""
            QPushButton {{
                background: transparent; color: {colors.TEXT_PRIMARY};
                padding: 8px 16px; border-radius: 8px; font-weight: 600;
                font-size: 13px; border: 1.5px solid {colors.BORDER};
            }}
            QPushButton:hover {{ background: {colors.BG_MAIN}; border-color: {colors.TEXT_SECONDARY}; }}
        """
        )
        btn_print.clicked.connect(self.print_student_list)

        btn_excel = QPushButton("📊  Excel")
        btn_excel.setMinimumHeight(42)
        btn_excel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_excel.setStyleSheet(
            f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {colors.SUCCESS}, stop:1 #16A34A);
                color: white; padding: 8px 16px; border-radius: 8px;
                font-weight: 600; font-size: 13px; border: none;
            }}
            QPushButton:hover {{ background: {colors.SUCCESS_HOVER}; }}
        """
        )
        btn_excel.clicked.connect(self._export_excel)

        _lbl_du = QLabel("De:")
        _lbl_du.setStyleSheet(f"color: {colors.TEXT_SECONDARY}; font-size: 12px; font-weight: 600;")
        _lbl_au = QLabel("À:")
        _lbl_au.setStyleSheet(f"color: {colors.TEXT_SECONDARY}; font-size: 12px; font-weight: 600;")
        flay.addWidget(self.combo_filter_cycle)
        flay.addWidget(self.combo_filter_class)
        flay.addWidget(_lbl_du)
        flay.addWidget(self.date_filter_from)
        flay.addWidget(_lbl_au)
        flay.addWidget(self.date_filter_to)
        flay.addWidget(self.txt_search, 1)
        flay.addWidget(btn_print)
        flay.addWidget(btn_excel)

        layout.addWidget(filter_frame)

        self.table_students = QTableWidget()
        style_table(self.table_students)
        self.table_students.setColumnCount(11)
        self.table_students.setHorizontalHeaderLabels(
            [
                "ID",
                "Prénom & Nom (FR)",
                "الاسم واللقب (AR)",
                "Sexe",
                "Classe",
                "N° Classe",
                "Code Accès",
                "Tuteur",
                "Tél",
                "Date",
                "Actions",
            ]
        )
        self.table_students.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_students.setColumnWidth(0, 50)
        self.table_students.setColumnWidth(3, 60)

        layout.addWidget(self.table_students)

        self.empty_state_list = EmptyStateWidget(
            icon="🎓",
            title="Aucun élève trouvé / لا يوجد طلاب",
            subtitle="Modifiez les filtres ou ajoutez un nouvel élève.",
        )
        self.empty_state_list.setVisible(False)
        layout.addWidget(self.empty_state_list)

        self.pagination_list = PaginationWidget(page_size=50)
        self.pagination_list.page_changed.connect(lambda _: self.populate_table(self.table_students))
        layout.addWidget(self.pagination_list)

        self.tabs.addTab(tab, "  Liste Complète / القائمة الشاملة  ")

    # ===== Logic Methods =====

    def get_active_year_id(self):
        try:
            db = DatabaseManager()
            with db.get_connection() as conn:
                return StudentRepository(conn).get_active_year_id() or None
        except Exception:
            return None

    def gender_to_index(self, value):
        if value in (None, "", 0, "0", "M", "m", "Masculin", "Male", "Homme"):
            return 0
        if value in (1, "1", "F", "f", "Féminin", "Feminin", "Female", "Femme"):
            return 1
        try:
            return 0 if int(value) == 0 else 1
        except (TypeError, ValueError):
            return 0

    def gender_to_label(self, value):
        return "M" if self.gender_to_index(value) == 0 else "F"

    def get_next_class_number(self, class_id, year_id):
        db = DatabaseManager()
        with db.get_connection() as conn:
            return StudentRepository(conn).get_next_class_number(class_id, year_id)

    def get_student_class_number(self, student_id, year_id):
        db = DatabaseManager()
        with db.get_connection() as conn:
            return StudentRepository(conn).get_class_assignment(student_id, year_id)

    def assign_class_number(self, student_id, class_id):
        year_id = self.get_active_year_id()
        if not year_id or not class_id:
            return None

        existing = self.get_student_class_number(student_id, year_id)
        if existing and existing[0] == class_id:
            return existing[1]

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = StudentRepository(conn)
            new_number = repo.get_next_class_number(class_id, year_id)
            repo.set_class_assignment(student_id, class_id, year_id, new_number)
            conn.commit()
            return new_number

    # ================== فتح Dialog الطالب ==================

    def _build_dialog(self, data=None):
        """ينشئ StudentDialog مع تمرير الدوال اللازمة."""
        # تحميل قائمة الدورات
        db = DatabaseManager()
        with db.get_connection() as conn:
            cycles = StudentRepository(conn).list_cycles()
        cycles_data = [(r[0], r[1]) for r in cycles]

        def get_classes(cycle_id):
            _db = DatabaseManager()
            with _db.get_connection() as conn:
                rows = StudentRepository(conn).list_classes(cycle_id)
            return [(r[0], r[1]) for r in rows]

        def get_next_number(class_id):
            year_id = self.get_active_year_id()
            if not year_id:
                return 1
            return self.get_next_class_number(class_id, year_id)

        return StudentDialog(cycles_data, get_classes, get_next_number, data=data, parent=self)

    def open_student_dialog(self, student_id=None):
        """يفتح dialog الإضافة (student_id=None) أو التعديل."""
        data = None
        if student_id is not None:
            data = self._load_student_dict(student_id)
            if data is None:
                QMessageBox.warning(self, "Erreur", "Impossible de charger les données de l'élève.")
                return

        dialog = self._build_dialog(data=data)
        result = dialog.exec()

        if result == QDialog.DialogCode.Accepted:
            values = dialog.get_values()
            if student_id is None:
                self._save_new_student(values)
            else:
                self._save_edit_student(student_id, values)
        elif result == StudentDialog.RESULT_SAVE_AND_NEW:
            # حفظ ثم فتح نموذج جديد فارغ
            values = dialog.get_values()
            self._save_new_student(values)
            self.open_student_dialog()  # إعادة فتح Dialog جديد

    def _load_student_dict(self, student_id) -> dict | None:
        """يحمّل بيانات الطالب من قاعدة البيانات ويحوّلها إلى dict."""
        try:
            active_year = self.get_active_year_id()
            db = DatabaseManager()
            with db.get_connection() as conn:
                repo = StudentRepository(conn)
                row = repo.get_student_for_edit(active_year, student_id)
                # نجلب رقم الفصل للعرض في الـ Dialog
                assignment = repo.get_class_assignment(student_id, active_year) if active_year else None
            if not row:
                return None
            return {
                "first_name_fr": row[0],
                "last_name_fr": row[1],
                "first_name_ar": row[2],
                "last_name_ar": row[3],
                "birth_date": row[4],
                "birth_place": row[5],
                "gender": row[6],
                "address": row[7],
                "parent_name": row[8],
                "parent_phone": row[9],
                "parent_email": row[10],
                "parent_address": row[11],
                "registration_date": row[12],
                "status": row[13],
                "photo_path": row[14],
                "class_id": row[15],
                "class_number": assignment[1] if assignment else None,
            }
        except Exception as e:
            AppLogger.error("StudentManagement", f"_load_student_dict error: {e}")
            return None

    def _save_new_student(self, values: dict):
        """يحفظ طالباً جديداً من قيم الـ Dialog."""
        try:
            photo_path = self._persist_photo(values.get("photo_path"), new_id=None)
            if photo_path:
                values["photo_path"] = photo_path

            db = DatabaseManager()
            with db.get_connection() as conn:
                repo = StudentRepository(conn)
                student_id = repo.add_student(values)
                year_id = repo.get_active_year_id()
                class_number = None
                if year_id and values.get("class_id"):
                    new_num = repo.get_next_class_number(values["class_id"], year_id)
                    repo.set_class_assignment(student_id, values["class_id"], year_id, new_num)
                    class_number = new_num
                from database_setup import log_audit

                log_audit(
                    conn,
                    getattr(self, "current_user", "system"),
                    "ADD_STUDENT",
                    f"{values['first_name_fr']} {values['last_name_fr']}",
                )
                conn.commit()

            if class_number is None:
                QMessageBox.warning(self, "Attention", "Aucune année scolaire active — رقم الفصل لم يُحدَّد.")
            else:
                ToastNotification.show_toast(self, f"Élève ajouté — N° Classe: {class_number}", kind="success")
            self.refresh_student_list()
        except Exception as e:
            QMessageBox.critical(self, "Erreur", friendly_db_error(e))

    def _save_edit_student(self, student_id: int, values: dict):
        """يحدّث بيانات طالب موجود من قيم الـ Dialog."""
        try:
            photo_path = self._persist_photo(values.get("photo_path"), new_id=student_id)
            if photo_path:
                values["photo_path"] = photo_path

            db = DatabaseManager()
            with db.get_connection() as conn:
                repo = StudentRepository(conn)
                repo.update_student(student_id, values)
                year_id = repo.get_active_year_id()
                if year_id and values.get("class_id"):
                    existing = repo.get_class_assignment(student_id, year_id)
                    if not existing or existing[0] != values["class_id"]:
                        new_num = repo.get_next_class_number(values["class_id"], year_id)
                        repo.set_class_assignment(student_id, values["class_id"], year_id, new_num)
                from database_setup import log_audit

                log_audit(
                    conn,
                    getattr(self, "current_user", "system"),
                    "EDIT_STUDENT",
                    f"{values['first_name_fr']} {values['last_name_fr']} (id={student_id})",
                )
                conn.commit()

            ToastNotification.show_toast(self, "Mise à jour réussie.", kind="success")
            self.refresh_student_list()
        except Exception as e:
            QMessageBox.critical(self, "Erreur", friendly_db_error(e))

    def _persist_photo(self, source_path: str | None, new_id) -> str | None:
        """يحفظ الصورة في school_data/photos إذا كانت مصدراً خارجياً."""
        if not source_path:
            return None
        if "school_data" in source_path:
            return source_path  # مسار داخلي موجود مسبقاً
        try:
            os.makedirs("school_data/photos", exist_ok=True)
            suffix = f"_{new_id}" if new_id else ""
            filename = f"student{suffix}_{datetime.now().timestamp()}.jpg"
            dest = f"school_data/photos/{filename}"
            shutil.copy(source_path, dest)
            return dest
        except Exception as e:
            AppLogger.error("StudentManagement", f"Photo copy error: {e}")
            return None

    # ================== add_student / save_student / update_student (legacy kept for compatibility) ==================
    def add_student(self):
        self.open_student_dialog()

    def save_student(self):
        if self.selected_student_id:
            self.open_student_dialog(self.selected_student_id)
        else:
            self.open_student_dialog()

    def update_student(self):
        if self.selected_student_id:
            self.open_student_dialog(self.selected_student_id)

    def delete_student(self, student_id):
        reply = QMessageBox.question(
            self,
            "Confirmation",
            "Supprimer cet étudiant ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            try:
                db = DatabaseManager()
                with db.get_connection() as conn:
                    repo = StudentRepository(conn)
                    repo.delete_student(student_id)
                    conn.commit()
                    from database_setup import log_audit

                    log_audit(conn, getattr(self, "current_user", "system"), "DELETE_STUDENT", f"id={student_id}")
                self.refresh_student_list()
            except Exception as e:
                QMessageBox.critical(self, "Erreur", friendly_db_error(e))

    def load_cycles_filter(self):
        self._load_cycles_into(self.combo_filter_cycle, "Tous les cycles")

    def _load_cycles_into(self, combo, default_text):
        db = DatabaseManager()
        with db.get_connection() as conn:
            rows = StudentRepository(conn).list_cycles()
        combo.clear()
        combo.addItem(default_text, None)
        for r in rows:
            combo.addItem(r[1], r[0])

    def load_classes_for_reg(self):
        cid = self.combo_cycle_reg.currentData()
        self._load_classes_into(self.combo_class_reg, cid, "Choisis Classe...")
        self.update_class_number_preview()

    def load_classes_for_filter(self):
        cid = self.combo_filter_cycle.currentData()
        self._load_classes_into(self.combo_filter_class, cid, "Toutes les classes")

    def _load_classes_into(self, combo, cycle_id, default_text):
        db = DatabaseManager()
        with db.get_connection() as conn:
            rows = StudentRepository(conn).list_classes(cycle_id)
        combo.clear()
        combo.addItem(default_text, None)
        for r in rows:
            combo.addItem(r[1], r[0])

    def refresh_student_list(self):
        if hasattr(self, "pagination_list"):
            self.pagination_list.reset()
        self.populate_table(self.table_students)
        self.populate_table(self.table_students_reg)
        self._load_kpi_stats()

    def populate_table(self, table):
        table.setRowCount(0)
        if table == self.table_students:
            cycle_id = self.combo_filter_cycle.currentData()
            class_id = self.combo_filter_class.currentData()
            search = self.txt_search.text().strip()
            date_from = (
                self.date_filter_from.date().toString("yyyy-MM-dd") if hasattr(self, "date_filter_from") else None
            )
            date_to = self.date_filter_to.date().toString("yyyy-MM-dd") if hasattr(self, "date_filter_to") else None
        else:
            cycle_id = None
            class_id = self.combo_filter_class_reg.currentData()
            search = self.txt_search_reg.text().strip()
            date_from = None
            date_to = None

        active_year_id = self.get_active_year_id()
        year_param = active_year_id or -1

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = StudentRepository(conn)
            if table == self.table_students and hasattr(self, "pagination_list"):
                total = repo.count_students(
                    year_param,
                    cycle_id=cycle_id,
                    class_id=class_id,
                    search=search,
                    date_from=date_from,
                    date_to=date_to,
                )
                rows = repo.list_students(
                    year_param,
                    cycle_id=cycle_id,
                    class_id=class_id,
                    search=search,
                    date_from=date_from,
                    date_to=date_to,
                    limit=self.pagination_list.page_size,
                    offset=self.pagination_list.current_offset(),
                )
                self.pagination_list.set_total(total)
                self.table_students.setVisible(total > 0)
                self.empty_state_list.setVisible(total == 0)
            else:
                rows = repo.list_students(
                    year_param,
                    cycle_id=cycle_id,
                    class_id=class_id,
                    search=search,
                    date_from=date_from,
                    date_to=date_to,
                )

        for r in rows:
            row_idx = table.rowCount()
            table.insertRow(row_idx)
            full_name_fr = f"{r[1]} {r[2]}"
            full_name_ar = f"{r[3]} {r[4]}"
            gender_str = self.gender_to_label(r[5])
            class_no = str(r[10]) if r[10] else "-"

            if table == self.table_students_reg:
                name_lang = self.combo_name_lang_reg.currentData() if hasattr(self, "combo_name_lang_reg") else "fr"
                if name_lang == "ar":
                    self.table_students_reg.setHorizontalHeaderLabels(
                        ["ID", "رقم الفصل", "الاسم الكامل", "الفصل", "⚙️ Actions"]
                    )
                    name_value = full_name_ar.strip()
                else:
                    self.table_students_reg.setHorizontalHeaderLabels(
                        ["ID", "N° Classe", "Prénom & Nom", "Classe", "⚙️ Actions"]
                    )
                    name_value = full_name_fr.strip()

                table.setItem(row_idx, 0, QTableWidgetItem(str(r[0])))
                table.setItem(row_idx, 1, QTableWidgetItem(class_no))
                name_item = QTableWidgetItem(name_value)
                if name_lang == "ar":
                    name_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                table.setItem(row_idx, 2, name_item)
                table.setItem(row_idx, 3, QTableWidgetItem(r[6] or "-"))  # class_name_fr
                table.setRowHeight(row_idx, 38)

                _c = ThemeManager.get_colors()
                container = QWidget()
                layout = QHBoxLayout(container)
                layout.setContentsMargins(4, 3, 4, 3)
                layout.setSpacing(6)

                btn_edit = QPushButton("✎ Modifier")
                btn_edit.setFixedHeight(28)
                btn_edit.setCursor(Qt.CursorShape.PointingHandCursor)
                btn_edit.setStyleSheet(
                    f"QPushButton {{ background:{_c.PRIMARY}; color:white; border-radius:6px;"
                    f"border:none; font-weight:600; font-size:11px; padding:3px 8px; }}"
                    f"QPushButton:hover {{ background:{_c.PRIMARY_HOVER}; }}"
                )
                btn_edit.clicked.connect(lambda ch, sid=r[0]: self.open_student_dialog(sid))
                btn_edit.setVisible(self._rbac_can_write)

                btn_del = QPushButton("✕ Suppr.")
                btn_del.setFixedHeight(28)
                btn_del.setCursor(Qt.CursorShape.PointingHandCursor)
                btn_del.setStyleSheet(
                    f"QPushButton {{ background:{_c.DANGER}; color:white; border-radius:6px;"
                    f"border:none; font-weight:600; font-size:11px; padding:3px 8px; }}"
                    f"QPushButton:hover {{ background:#B91C1C; }}"
                )
                btn_del.clicked.connect(lambda ch, sid=r[0]: self.delete_student(sid))
                btn_del.setVisible(self._rbac_can_delete)

                layout.addWidget(btn_edit)
                layout.addWidget(btn_del)
                table.setCellWidget(row_idx, 4, container)

            else:  # Full Table
                table.setItem(row_idx, 0, QTableWidgetItem(str(r[0])))
                full_fr = f"{r[1]} {r[2]}".strip()
                full_ar = f"{r[3]} {r[4]}".strip()
                table.setItem(row_idx, 1, QTableWidgetItem(full_fr))
                table.setItem(row_idx, 2, QTableWidgetItem(full_ar))
                table.setItem(row_idx, 3, QTableWidgetItem(gender_str))
                table.setItem(row_idx, 4, QTableWidgetItem(r[6] or "-"))
                table.setItem(row_idx, 5, QTableWidgetItem(class_no))
                table.setItem(row_idx, 6, QTableWidgetItem(r[11] or "-"))  # student_code
                table.setItem(row_idx, 7, QTableWidgetItem(r[7] or ""))
                table.setItem(row_idx, 8, QTableWidgetItem(r[8] or ""))
                table.setItem(row_idx, 9, QTableWidgetItem(str(r[9]) if r[9] else ""))

                container = QWidget()
                layout = QHBoxLayout(container)
                layout.setContentsMargins(2, 2, 2, 2)
                _c2 = ThemeManager.get_colors()
                btn_del = QPushButton("✕")
                btn_del.setFixedSize(24, 24)
                btn_del.setStyleSheet(f"background: {_c2.DANGER}; color: white; border-radius: 4px; border: none;")
                btn_del.clicked.connect(lambda ch, sid=r[0]: self.delete_student(sid))
                btn_del.setVisible(self._rbac_can_delete)
                layout.addWidget(btn_del)
                table.setCellWidget(row_idx, 10, container)

    def print_student_list(self):
        rows = self._fetch_full_list_rows()
        headers = [
            "ID",
            "Prénom (FR)",
            "Nom (FR)",
            "الاسم (AR)",
            "اللقب (AR)",
            "Naissance",
            "Lieu",
            "Sexe",
            "Adresse",
            "Classe",
            "N° Cls",
            "Code Accès",
            "Tuteur",
            "Tél",
            "Email",
            "Date",
            "Statut",
        ]
        filename = self._build_report_filename("Liste_Complete_Eleves", self.combo_filter_class.currentText())
        self._generate_pdf_rows(
            headers,
            rows,
            filename,
            orientation='L',
            title_doc="LISTE COMPLETE DES ELEVES INSCRITES DE: "
            + (
                self.combo_filter_class.currentText() if self.combo_filter_class.currentData() else "Toutes les classes"
            ),
        )

    def _open_import_wizard(self):
        from import_wizard import ImportWizard

        actor = getattr(self, "current_user", "system")
        wiz = ImportWizard(parent=self, actor=actor)
        wiz.exec()
        self.refresh_student_list()

    # ───────────────────────────── KPI ─────────────────────────────

    def _load_kpi_stats(self):
        """Charge les compteurs rapides pour les cartes KPI."""
        try:
            db = DatabaseManager()
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM Students")
                total = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM Students WHERE status = 'Active'")
                actifs = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM Students WHERE gender = 'M'")
                garcons = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM Students WHERE gender = 'F'")
                filles = cursor.fetchone()[0]
            self._stat_total.set_value(str(total))
            self._stat_actifs.set_value(str(actifs))
            self._stat_garcons.set_value(str(garcons))
            self._stat_filles.set_value(str(filles))
        except Exception as e:
            AppLogger.error("StudentManagement", f"KPI load error: {e}")

    # ─────────────────────────── EXPORT EXCEL ───────────────────────────

    def _export_excel(self):
        """Exporte la liste filtrée vers un fichier Excel (.xlsx)."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            QMessageBox.critical(self, "Erreur", "openpyxl n'est pas installé.")
            return

        rows = self._fetch_full_list_rows()
        if not rows:
            QMessageBox.information(self, "Export", "Aucune donnée à exporter.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer Excel", f"Eleves_{datetime.now().strftime('%Y%m%d')}.xlsx", "Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Élèves"

            headers = [
                "ID",
                "Prénom (FR)",
                "Nom (FR)",
                "الاسم (AR)",
                "اللقب (AR)",
                "Naissance",
                "Lieu",
                "Sexe",
                "Adresse",
                "Classe",
                "N° Cls",
                "Code Accès",
                "Tuteur",
                "Tél",
                "Email",
                "Date Inscr.",
                "Statut",
            ]
            header_fill = PatternFill("solid", fgColor="1E293B")
            header_font = Font(bold=True, color="F8FAFC")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_data in rows:
                ws.append(list(row_data))

            # Ajustement automatique de la largeur des colonnes
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(file_path)
            ToastNotification.show_toast(self, f"Excel exporté : {file_path}", kind="success")
        except Exception as e:
            QMessageBox.critical(self, "Erreur Export", str(e))

    def print_filtered_list(self):
        rows, headers = self._fetch_reg_list_rows()
        filename = self._build_report_filename("Liste_Eleves_Filtrees", self.combo_filter_class_reg.currentText())
        self._generate_pdf_rows(
            headers,
            rows,
            filename,
            orientation='P',
            title_doc="LISTE DES ELEVES DE: "
            + (
                self.combo_filter_class_reg.currentText()
                if self.combo_filter_class_reg.currentData()
                else "Toutes les classes"
            ),
        )

    def _build_report_filename(self, prefix, class_label):
        class_text = class_label if class_label and "Toutes" not in class_label else "Toutes_Classes"
        class_slug = "".join(ch for ch in class_text.replace(" ", "_") if ch.isalnum() or ch in "-_") or "Classe"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{class_slug}_{timestamp}.pdf"

    def _get_school_info(self):
        try:
            db = DatabaseManager()
            with db.get_connection() as conn:
                return FinanceRepository(conn).get_school_info()
        except Exception:
            return None

    def _fetch_full_list_rows(self):
        cycle_id = self.combo_filter_cycle.currentData()
        class_id = self.combo_filter_class.currentData()
        search = self.txt_search.text().strip()
        date_from = self.date_filter_from.date().toString("yyyy-MM-dd") if hasattr(self, "date_filter_from") else None
        date_to = self.date_filter_to.date().toString("yyyy-MM-dd") if hasattr(self, "date_filter_to") else None
        active_year_id = self.get_active_year_id()
        year_param = active_year_id or -1

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = StudentRepository(conn)
            rows = repo.list_students_detailed(
                year_param,
                cycle_id=cycle_id,
                class_id=class_id,
                search=search,
                date_from=date_from,
                date_to=date_to,
            )

        result = []
        for r in rows:
            gender_str = self.gender_to_label(r[7])
            class_no = str(r[10]) if r[10] else "-"
            result.append(
                [
                    str(r[0]),
                    r[1] or "",
                    r[2] or "",
                    r[3] or "",
                    r[4] or "",
                    r[5] or "",
                    r[6] or "",
                    gender_str,
                    r[8] or "",
                    r[9] or "-",
                    class_no,
                    r[11] or "",
                    r[12] or "",
                    r[13] or "",
                    r[14] or "",
                    r[15] or "",
                    r[16] or "",
                ]
            )
        return result

    def _fetch_reg_list_rows(self):
        name_lang = self.combo_name_lang_reg.currentData() if hasattr(self, "combo_name_lang_reg") else "fr"
        headers = ["N° Classe", "Nom & Prénom"] if name_lang == "fr" else ["رقم الفصل", "الاسم الكامل"]
        rows = []
        for row in range(self.table_students_reg.rowCount()):
            class_item = self.table_students_reg.item(row, 1)
            name_item = self.table_students_reg.item(row, 2)
            class_val = class_item.text() if class_item else ""
            name_val = name_item.text() if name_item else ""
            rows.append([class_val, name_val])
        return rows, headers

    def _generate_pdf_rows(self, headers, rows, filename, orientation='P', title_doc="LISTE DES ELEVES INSCRITES"):
        pdf = StudentListPDF(self._get_school_info(), title_doc=title_doc, orientation=orientation)
        self._setup_pdf_fonts(pdf)
        pdf.add_page()
        pdf.set_font(self._get_pdf_font_name(), size=8 if orientation == 'L' else 10)

        col_count = len(headers)
        page_w = pdf.w - 20
        col_widths = None
        if orientation == 'L' and col_count == 16:
            col_widths = [10, 20, 15, 18, 13, 16, 17, 10, 24, 13, 12, 33, 16, 35, 15, 10]
        if orientation == 'L' and col_count == 17:
            col_widths = [9, 18, 14, 16, 12, 16, 14, 9, 22, 12, 12, 22, 15, 35, 20, 14, 12]
        w = page_w / max(1, col_count)

        pdf.set_draw_color(203, 213, 225)
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(248, 250, 252)
        for i, h in enumerate(headers):
            txt = self._prepare_pdf_text(h)
            cell_w = col_widths[i] if col_widths else w
            pdf.cell(cell_w, 9, txt, 1, 0, 'C', fill=True)
        pdf.ln()

        pdf.set_text_color(51, 65, 85)
        row_fill = False

        for row in rows:
            pdf.set_fill_color(241, 245, 249) if row_fill else pdf.set_fill_color(255, 255, 255)
            for col in range(col_count):
                txt = row[col] if col < len(row) else ""
                txt = self._prepare_pdf_text(txt)
                try:
                    txt.encode('latin-1')
                except Exception:
                    if not self._is_arabic_font_ready():
                        txt = self._latin_fallback_text(txt)
                cell_w = col_widths[col] if col_widths else w
                pdf.cell(cell_w, 8, txt, 1, 0, fill=row_fill)
            pdf.ln()
            row_fill = not row_fill

        output_pdf(
            pdf,
            self,
            filename,
            mode=STUDENT_LIST_OUTPUT_MODE,
            dialog_title="Save PDF",
            success_save_message="PDF généré.",
            success_print_message="Liste envoyée à l'imprimante.",
        )

    def _generate_pdf(self, table, filename):
        pdf = StudentListPDF()
        self._setup_pdf_fonts(pdf)
        pdf.add_page()
        pdf.set_font(self._get_pdf_font_name(), size=10)

        col_count = table.columnCount()
        w = 190 / max(1, col_count)
        pdf.set_draw_color(203, 213, 225)
        pdf.set_text_color(51, 65, 85)
        row_fill = False

        for row in range(table.rowCount()):
            pdf.set_fill_color(241, 245, 249) if row_fill else pdf.set_fill_color(255, 255, 255)
            for col in range(col_count):
                item = table.item(row, col)
                txt = item.text() if item else ""
                txt = txt.replace("\n", " / ")
                txt = self._prepare_pdf_text(txt)
                try:
                    txt.encode('latin-1')
                except Exception:
                    if not self._is_arabic_font_ready():
                        txt = self._latin_fallback_text(txt)
                pdf.cell(w, 10, txt, 1, 0, fill=row_fill)
            pdf.ln()
            row_fill = not row_fill

        output_pdf(
            pdf,
            self,
            filename,
            mode=STUDENT_LIST_OUTPUT_MODE,
            dialog_title="Save PDF",
            success_save_message="PDF généré.",
            success_print_message="Liste envoyée à l'imprimante.",
        )

    def _get_pdf_font_name(self):
        return "ArabicFont" if self._is_arabic_font_ready() else "Arial"

    def _latin_fallback_text(self, text):
        return latin_fallback(text)

    def _is_arabic_font_ready(self):
        return is_arabic_font_ready()

    def _setup_pdf_fonts(self, pdf):
        if not setup_pdf_arabic_font(pdf) and ARABIC_SUPPORT:
            QMessageBox.information(
                self, "Police عربية مفقودة", "لتصدير PDF يدعم العربية، ضع ملف خط (TTF) في مجلد fonts داخل المشروع."
            )

    def _prepare_pdf_text(self, text):
        return prepare_pdf_text(text)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ModernStudentManagement()
    window.show()
    sys.exit(app.exec())
