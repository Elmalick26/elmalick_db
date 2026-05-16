import sys
import os
import psycopg2
from datetime import datetime, timedelta
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QComboBox,
                             QMessageBox, QGroupBox, QTabWidget, QFrame,
                             QGridLayout, QGraphicsDropShadowEffect, QDateEdit,
                             QFileDialog, QProgressBar)
from PyQt6.QtCore import Qt, QDate, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
from database_setup import DatabaseManager
from repositories.analytics_repo import AnalyticsRepository
from repositories.finance_repo import FinanceRepository
from pdf_report_style import apply_grades_sheet_header, apply_table_header_style, apply_table_body_style, set_zebra_row_fill, get_school_info_row
from print_export_service import output_pdf, get_report_output_mode
from app_logger import AppLogger

from ui_styles import ThemeManager, get_card_style, apply_shadow_to_widget, Colors, get_tabs_style

THEME_AVAILABLE = True
ADVANCED_COMPREHENSIVE_PDF_MODE = get_report_output_mode("advanced_comprehensive_pdf_mode", "save")

# --- Worker Thread for Heavy Report Generation ---


class ReportWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)  # Success message or file path
    error = pyqtSignal(str)

    def __init__(self, report_type, params):
        super().__init__()
        self.report_type = report_type
        self.params = params

    # ===== تم التعديل: جلب أعمدة الجدول بالطريقة القياسية لـ PostgreSQL =====
    def _student_name_sql_expr(self, columns: set, alias='s'):
        parts = []
        if {'first_name_fr', 'last_name_fr'}.issubset(columns):
            parts.append(f"TRIM(NULLIF({alias}.first_name_fr, '') || ' ' || NULLIF({alias}.last_name_fr, ''))")
        if 'student_name' in columns:
            parts.append(f"NULLIF({alias}.student_name, '')")
        if 'first_name_fr' in columns:
            parts.append(f"NULLIF({alias}.first_name_fr, '')")
        if 'last_name_fr' in columns:
            parts.append(f"NULLIF({alias}.last_name_fr, '')")

        if not parts:
            return "'N/A'"
        return f"COALESCE({', '.join(parts)}, 'N/A')"

    def run(self):
        try:
            if self.report_type == "excel_financial":
                result = self.generate_financial_excel()
            elif self.report_type == "excel_students":
                result = self.generate_students_excel()
            elif self.report_type == "excel_attendance":
                result = self.generate_attendance_excel()
            elif self.report_type == "excel_grades":
                result = self.generate_grades_excel()
            else:
                result = "Type de rapport inconnu"

            self.finished.emit(result)
        except Exception as e:
            AppLogger.error("ReportWorker", f"Erreur lors de la génération du rapport {self.report_type}", e)
            self.error.emit(str(e))

    def generate_financial_excel(self):
        """Generate comprehensive financial Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport Financier"
        selected_period = self.params.get("period", "12 derniers mois") if self.params else "12 derniers mois"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        ws['A1'] = "RAPPORT FINANCIER COMPLET"
        ws['A1'].font = Font(bold=True, size=16, color="1E293B")
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:F2')
        ws['A3'] = f"Période: {selected_period}"
        ws.merge_cells('A3:F3')

        self.progress.emit(20)

        ws['A4'] = "RECETTES (Paiements Étudiants)"
        ws['A4'].font = Font(bold=True, size=14)

        ws['A5'] = "Mois"
        ws['B5'] = "Nb Paiements"
        ws['C5'] = "Montant Total"
        for cell in ['A5', 'B5', 'C5']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            income_data = repo.get_monthly_income_totals(selected_period)
            expense_data = repo.get_monthly_expense_totals(selected_period)

            row = 6
            total_income = 0
            for month, count, total in income_data:
                ws[f'A{row}'] = month
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"{total:,.0f} FCFA"
                total_income += total if total else 0
                row += 1

            ws[f'A{row}'] = "TOTAL RECETTES"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'] = f"{total_income:,.0f} FCFA"
            ws[f'C{row}'].font = Font(bold=True, color="10B981")

            self.progress.emit(50)

            ws[f'A{row + 2}'] = "DÉPENSES"
            ws[f'A{row + 2}'].font = Font(bold=True, size=14)

            header_row = row + 3
            ws[f'A{header_row}'] = "Mois"
            ws[f'B{header_row}'] = "Nb Dépenses"
            ws[f'C{header_row}'] = "Montant Total"
            for cell in [f'A{header_row}', f'B{header_row}', f'C{header_row}']:
                ws[cell].fill = header_fill
                ws[cell].font = header_font

            exp_row = header_row + 1
            total_expenses = 0
            for month, count, total in expense_data:
                ws[f'A{exp_row}'] = month
                ws[f'B{exp_row}'] = count
                ws[f'C{exp_row}'] = f"{total:,.0f} FCFA"
                total_expenses += total if total else 0
                exp_row += 1

            ws[f'A{exp_row}'] = "TOTAL DÉPENSES"
            ws[f'A{exp_row}'].font = Font(bold=True)
            ws[f'C{exp_row}'] = f"{total_expenses:,.0f} FCFA"
            ws[f'C{exp_row}'].font = Font(bold=True, color="EF4444")

            self.progress.emit(80)

            balance = total_income - total_expenses
            ws[f'A{exp_row + 2}'] = "SOLDE NET"
            ws[f'A{exp_row + 2}'].font = Font(bold=True, size=14)
            ws[f'C{exp_row + 2}'] = f"{balance:,.0f} FCFA"
            ws[f'C{exp_row + 2}'].font = Font(bold=True, size=14, color="10B981" if balance >= 0 else "EF4444")

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20

        safe_period = selected_period.encode('ascii', 'ignore').decode('ascii')
        safe_period = safe_period.replace(" ", "_").replace("'", "").replace("/", "-")
        if not safe_period: safe_period = "periode"
        filename = f"Rapport_Financier_{safe_period}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
        filepath = self.params.get("output_path") if self.params else None
        if not filepath:
            filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)

        self.progress.emit(100)
        return filepath

    def generate_students_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Liste Étudiants"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            active_year_id, active_year_label = repo.get_active_year_context()
            student_cols = repo.get_student_columns()
            student_name_expr = self._student_name_sql_expr(student_cols, alias='s')

        ws['A1'] = "LISTE DES ÉTUDIANTS"
        ws['A1'].font = Font(bold=True, size=16, color="1E293B")
        ws.merge_cells('A1:L1')
        ws['A2'] = f"Année scolaire: {active_year_label}"
        ws.merge_cells('A2:L2')

        headers = [
            "ID", "N° Classe", "Nom Complet", "Classe", "Sexe",
            "Date Naissance", "Lieu Naissance", "Parent/Tuteur",
            "Téléphone", "Adresse", "Date Inscription", "Statut"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        self.progress.emit(30)

        birth_place_expr = "s.birth_place" if "birth_place" in student_cols else "''"
        parent_expr = "s.parent_name" if "parent_name" in student_cols else "''"
        phone_expr = "s.phone" if "phone" in student_cols else "''"
        address_expr = "s.address" if "address" in student_cols else "''"
        enroll_expr = "s.registration_date" if "registration_date" in student_cols else "''"

        with db.get_connection() as conn:
            cursor = conn.cursor()
            if active_year_id:
                cursor.execute(f"""
                    SELECT s.id,
                        COALESCE(CAST(scn.class_number AS VARCHAR), ''),
                        {student_name_expr} as name,
                        c.class_name_fr,
                        CASE WHEN s.gender='0' OR s.gender='M' THEN 'M' ELSE 'F' END as gender,
                        CAST(s.birth_date AS VARCHAR),
                        {birth_place_expr} as birth_place,
                        {parent_expr} as parent_name,
                        {phone_expr} as phone,
                        {address_expr} as address,
                        CAST({enroll_expr} AS VARCHAR) as registration_date,
                        s.status
                    FROM Students s
                    JOIN StudentClassNumbers scn ON scn.student_id = s.id AND scn.year_id = %s
                    JOIN Classes c ON scn.class_id = c.id
                    WHERE s.status='Active'
                    ORDER BY c.sort_order, COALESCE(scn.class_number, 9999), name
                """, (active_year_id,))
            else:
                cursor.execute(f"""
                    SELECT s.id,
                        '',
                        {student_name_expr} as name,
                        'N/A',
                        CASE WHEN s.gender='0' OR s.gender='M' THEN 'M' ELSE 'F' END as gender,
                        CAST(s.birth_date AS VARCHAR),
                        {birth_place_expr} as birth_place,
                        {parent_expr} as parent_name,
                        {phone_expr} as phone,
                        {address_expr} as address,
                        CAST({enroll_expr} AS VARCHAR) as registration_date,
                        s.status
                    FROM Students s
                    WHERE s.status='Active'
                    ORDER BY name
                """)
            rows = cursor.fetchall()

        for row, student in enumerate(rows, 5):
            for col, value in enumerate(student, 1):
                ws.cell(row=row, column=col, value=value)

            if row % 20 == 0:
                self.progress.emit(30 + (row * 60 // max(1, len(rows))))

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 22
        ws.column_dimensions['I'].width = 16
        ws.column_dimensions['J'].width = 24
        ws.column_dimensions['K'].width = 16
        ws.column_dimensions['L'].width = 12

        filename = f"Liste_Etudiants_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
        filepath = self.params.get("output_path") if self.params else None
        if not filepath:
            filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)

        self.progress.emit(100)
        return filepath

    def generate_attendance_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Assiduite"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws['A1'] = "RAPPORT D'ASSIDUITÉ"
        ws['A1'].font = Font(bold=True, size=16, color="1E293B")
        ws.merge_cells('A1:G1')
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:G2')

        headers = ["Classe", "Nb Étudiants", "Présents", "Absents", "Retards", "Total Pointages", "Taux Absence %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        self.progress.emit(20)

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            active_year_id, active_year_label = repo.get_active_year_context()

            ws['A3'] = f"Année scolaire: {active_year_label}"
            ws.merge_cells('A3:G3')

            rows = repo.get_attendance_summary_by_class(active_year_id) if active_year_id else []

        row_idx = 5
        for idx, row in enumerate(rows, 1):
            class_name, students_count, presents, absents, lates, attendance_rows = row
            presents = presents or 0
            absents = absents or 0
            lates = lates or 0
            attendance_rows = attendance_rows or 0
            absence_rate = (absents / attendance_rows * 100) if attendance_rows > 0 else 0

            ws.cell(row=row_idx, column=1, value=class_name or "-")
            ws.cell(row=row_idx, column=2, value=students_count or 0)
            ws.cell(row=row_idx, column=3, value=presents)
            ws.cell(row=row_idx, column=4, value=absents)
            ws.cell(row=row_idx, column=5, value=lates)
            ws.cell(row=row_idx, column=6, value=attendance_rows)
            ws.cell(row=row_idx, column=7, value=round(absence_rate, 2))
            row_idx += 1

            if idx % 5 == 0:
                self.progress.emit(min(90, 20 + idx * 5))

        for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], [24, 15, 12, 12, 12, 16, 14]):
            ws.column_dimensions[col].width = width

        filename = f"Rapport_Assiduite_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
        filepath = self.params.get("output_path") if self.params else None
        if not filepath:
            filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)

        self.progress.emit(100)
        return filepath

    def generate_grades_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Notes"

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws['A1'] = "RAPPORT GLOBAL DES NOTES"
        ws['A1'].font = Font(bold=True, size=16, color="1E293B")
        ws.merge_cells('A1:F1')
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:F2')

        headers = ["Classe", "Nb Étudiants", "Nb Notes", "Min", "Moyenne", "Max"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        self.progress.emit(20)

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            active_year_id, active_year_label = repo.get_active_year_context()

            ws['A3'] = f"Année scolaire: {active_year_label}"
            ws.merge_cells('A3:F3')

            rows = repo.get_grades_summary_by_class(active_year_id) if active_year_id else []

        row_idx = 5
        for idx, row in enumerate(rows, 1):
            class_name, students_count, grades_count, min_score, avg_score, max_score = row
            ws.cell(row=row_idx, column=1, value=class_name or "-")
            ws.cell(row=row_idx, column=2, value=students_count or 0)
            ws.cell(row=row_idx, column=3, value=grades_count or 0)
            ws.cell(row=row_idx, column=4, value=round(min_score, 2) if min_score is not None else "-")
            ws.cell(row=row_idx, column=5, value=round(avg_score, 2) if avg_score is not None else "-")
            ws.cell(row=row_idx, column=6, value=round(max_score, 2) if max_score is not None else "-")
            row_idx += 1

            if idx % 5 == 0:
                self.progress.emit(min(90, 20 + idx * 5))

        for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F'], [24, 15, 12, 10, 12, 10]):
            ws.column_dimensions[col].width = width

        filename = f"Rapport_Notes_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
        filepath = self.params.get("output_path") if self.params else None
        if not filepath:
            filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)

        self.progress.emit(100)
        return filepath


class AdvancedReportsWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Rapports Avancés / التقارير المتقدمة")
        self.setMinimumSize(1100, 700)

        if THEME_AVAILABLE:
            ThemeManager.apply_theme(self)
        else:
            colors = Colors()
            self.setStyleSheet(f"""
                QMainWindow {{ background-color: {colors.BG_MAIN}; }}
                QLabel {{ font-family: 'Segoe UI', 'Cairo', sans-serif; color: {colors.TEXT_PRIMARY}; }}
                QGroupBox {{
                    border: 1px solid {colors.BORDER}; border-radius: 8px; margin-top: 10px;
                    background-color: {colors.BG_CARD}; font-weight: bold; color: {colors.TEXT_SECONDARY};
                }}
            """)

        self.worker = None
        self.init_ui()

    def _sanitize_latin(self, text):
        if text is None:
            return ""
        if not isinstance(text, str):
            text = str(text)
        return text.encode('latin-1', 'ignore').decode('latin-1')

    def _build_comprehensive_pdf(self, selected_period):
        pdf = FPDF()
        pdf.add_page()

        school_info = get_school_info_row()
        apply_grades_sheet_header(pdf, school_info, "RAPPORT COMPREHENSIF")

        period_text = self._sanitize_latin(
            f"Période: {selected_period} | Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        )
        pdf.set_font("Arial", '', 10)
        pdf.set_text_color(51, 65, 85)
        pdf.cell(0, 7, period_text, 0, 1, 'C')
        pdf.ln(2)

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            active_year_id, active_year_label = repo.get_active_year_context()
            total_income = repo.get_total_income_by_period(selected_period)
            total_expenses = repo.get_total_expense_by_period(selected_period)
            balance = total_income - total_expenses

            if active_year_id:
                stats = AnalyticsRepository(conn).get_comprehensive_stats(active_year_id)
            else:
                stats = AnalyticsRepository(conn).get_comprehensive_stats(None)
            active_students = stats["active_students"]
            total_classes = stats["total_classes"]
            presents = stats["presents"]
            absents = stats["absents"]
            lates = stats["lates"]

        apply_table_header_style(pdf, "Arial", 10)
        pdf.cell(0, 8, self._sanitize_latin("Résumé Financier"), 1, 1, 'L', True)
        apply_table_body_style(pdf, "Arial", 10)

        fin_rows = [
            ("Total Recettes", f"{total_income:,.0f} FCFA"),
            ("Total Dépenses", f"{total_expenses:,.0f} FCFA"),
            ("Solde Net", f"{balance:,.0f} FCFA"),
        ]
        for idx, (label, value) in enumerate(fin_rows):
            set_zebra_row_fill(pdf, idx)
            pdf.cell(110, 8, self._sanitize_latin(label), 1, 0, 'L', True)
            pdf.cell(80, 8, self._sanitize_latin(value), 1, 1, 'R', True)

        pdf.ln(3)
        apply_table_header_style(pdf, "Arial", 10)
        pdf.cell(0, 8, self._sanitize_latin("Résumé Académique"), 1, 1, 'L', True)
        apply_table_body_style(pdf, "Arial", 10)

        acad_rows = [
            ("Année scolaire active", active_year_label),
            ("Classes", str(total_classes)),
            ("Étudiants actifs", str(active_students)),
            ("Présences", str(presents)),
            ("Absences", str(absents)),
            ("Retards", str(lates)),
        ]
        for idx, (label, value) in enumerate(acad_rows):
            set_zebra_row_fill(pdf, idx)
            pdf.cell(110, 8, self._sanitize_latin(label), 1, 0, 'L', True)
            pdf.cell(80, 8, self._sanitize_latin(value), 1, 1, 'R', True)

        return pdf

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Header
        header_frame = QFrame()
        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()

        header_frame.setStyleSheet(f"""
            QFrame {{ background-color: {colors.BG_HEADER}; border-radius: 12px; }}
        """)
        header_frame.setFixedHeight(80)

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15); shadow.setColor(QColor(15, 23, 42, 40)); shadow.setOffset(0, 4)
        header_frame.setGraphicsEffect(shadow)

        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(20, 10, 20, 10)

        icon_lbl = QLabel("📊")
        icon_lbl.setStyleSheet("font-size: 32px; background: transparent;")

        title_box = QVBoxLayout()
        title = QLabel("RAPPORTS AVANCÉS & ANALYSES")
        title.setStyleSheet(f"color: {colors.HEADER_TEXT}; font-size: 20px; font-weight: bold; background: transparent;")

        subtitle = QLabel("Visualisations, Graphiques & Export Excel")
        subtitle.setStyleSheet(f"color: {colors.TEXT_SECONDARY}; font-size: 14px; background: transparent;")

        title_box.addWidget(title)
        title_box.addWidget(subtitle)

        header_layout.addWidget(icon_lbl)
        header_layout.addSpacing(15)
        header_layout.addLayout(title_box)
        header_layout.addStretch()

        layout.addWidget(header_frame)

        # Tabs
        tabs = QTabWidget()
        if THEME_AVAILABLE:
            tabs.setStyleSheet(get_tabs_style())

        tabs.addTab(self.create_financial_charts_tab(), "📊 Graphiques Financiers")
        tabs.addTab(self.create_student_reports_tab(), "📈 Performance Étudiants")
        tabs.addTab(self.create_excel_exports_tab(), "📑 Export Excel")

        layout.addWidget(tabs)

    def create_card(self):
        frame = QFrame()
        if THEME_AVAILABLE:
            frame.setStyleSheet(get_card_style())
            apply_shadow_to_widget(frame)
        else:
            colors = Colors()
            frame.setStyleSheet(f"QFrame {{ background-color: {colors.BG_CARD}; border-radius: 12px; border: 1px solid {colors.BORDER}; }}")
        return frame

    def styled_combo(self):
        combo = QComboBox()
        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()
        combo.setStyleSheet(f"""
            QComboBox {{ padding: 6px 12px; border: 1px solid {colors.BORDER}; border-radius: 6px; background: {colors.INPUT_BG}; color: {colors.TEXT_PRIMARY}; }}
            QComboBox:focus {{ border: 2px solid {colors.BORDER_FOCUS}; background: {colors.INPUT_BG_FOCUS}; }}
        """)
        combo.setMinimumHeight(38)
        return combo

    def create_financial_charts_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        control_card = self.create_card()
        controls_layout = QHBoxLayout(control_card)
        controls_layout.setContentsMargins(15, 15, 15, 15)

        controls_layout.addWidget(QLabel("Période:"))
        self.period_combo = self.styled_combo()
        self.period_combo.addItems(["6 derniers mois", "12 derniers mois", "Année en cours"])
        controls_layout.addWidget(self.period_combo)

        btn_generate = QPushButton("🔄 Actualiser Graphique")
        btn_generate.setCursor(Qt.CursorShape.PointingHandCursor)
        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()
        btn_generate.setStyleSheet(f"""
            QPushButton {{ background-color: {colors.PRIMARY}; color: white; padding: 8px 20px; border-radius: 6px; font-weight: bold; border: none; }}
            QPushButton:hover {{ background-color: {colors.PRIMARY_HOVER}; }}
        """)
        btn_generate.clicked.connect(self.generate_financial_chart)
        controls_layout.addWidget(btn_generate)
        controls_layout.addStretch()

        layout.addWidget(control_card)

        chart_card = self.create_card()
        c_layout = QVBoxLayout(chart_card)

        self.financial_figure = Figure(figsize=(10, 6))
        self.financial_figure.patch.set_facecolor(colors.BG_CARD)
        self.financial_canvas = FigureCanvas(self.financial_figure)
        c_layout.addWidget(self.financial_canvas)

        layout.addWidget(chart_card)
        self.generate_financial_chart()

        return widget

    def generate_financial_chart(self):
        self.financial_figure.clear()
        ax = self.financial_figure.add_subplot(111)
        selected_period = self.period_combo.currentText() if hasattr(self, 'period_combo') else "12 derniers mois"

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            income_data = repo.get_monthly_income_totals(selected_period)
            expense_data = repo.get_monthly_expense_totals(selected_period)

        income_dict = {item[0]: item[1] if item[1] else 0 for item in income_data}
        expense_dict = {item[0]: item[1] if item[1] else 0 for item in expense_data}
        months = sorted(set(income_dict.keys()) | set(expense_dict.keys()))
        income = [income_dict.get(month, 0) for month in months]
        expenses = [expense_dict.get(month, 0) for month in months]

        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()

        if not months:
            ax.text(0.5, 0.5, 'Aucune donnée disponible', ha='center', va='center', fontsize=14, color=colors.TEXT_SECONDARY)
            ax.axis('off')
        else:
            ax.set_facecolor(colors.BG_CARD)
            x = range(len(months))
            width = 0.35

            ax.bar([i - width / 2 for i in x], income, width, label='Recettes', color=colors.SUCCESS)
            ax.bar([i + width / 2 for i in x], expenses, width, label='Dépenses', color=colors.DANGER)

            ax.set_xlabel('Mois', fontsize=10, fontweight='bold', color=colors.TEXT_SECONDARY)
            ax.set_ylabel('Montant (FCFA)', fontsize=10, fontweight='bold', color=colors.TEXT_SECONDARY)
            ax.set_title('Évolution Financière Mensuelle', fontsize=12, fontweight='bold', pad=15, color=colors.TEXT_PRIMARY)
            ax.set_xticks(x)
            ax.set_xticklabels(months, rotation=45, ha='right')
            legend = ax.legend(loc='upper left', frameon=True)
            ax.grid(axis='y', alpha=0.2, linestyle='--')
            ax.tick_params(axis='x', colors=colors.TEXT_SECONDARY)
            ax.tick_params(axis='y', colors=colors.TEXT_SECONDARY)
            for spine in ax.spines.values(): spine.set_color(colors.BORDER)
            if legend:
                for text in legend.get_texts(): text.set_color(colors.TEXT_SECONDARY)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

        self.financial_figure.tight_layout()
        self.financial_canvas.draw()

    def create_student_reports_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        control_card = self.create_card()
        controls_layout = QHBoxLayout(control_card)
        controls_layout.setContentsMargins(15, 15, 15, 15)

        controls_layout.addWidget(QLabel("Classe:"))
        self.class_combo = self.styled_combo()
        self.load_classes_combo()
        controls_layout.addWidget(self.class_combo)

        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()
        btn_analyze = QPushButton("📈 Analyser Performance")
        btn_analyze.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_analyze.setStyleSheet(f"""
            QPushButton {{ background-color: {colors.SUCCESS}; color: white; padding: 8px 20px; border-radius: 6px; font-weight: bold; border: none; }}
            QPushButton:hover {{ background-color: {colors.SUCCESS_HOVER}; }}
        """)
        btn_analyze.clicked.connect(self.generate_student_chart)
        controls_layout.addWidget(btn_analyze)
        controls_layout.addStretch()
        layout.addWidget(control_card)

        chart_card = self.create_card()
        c_layout = QVBoxLayout(chart_card)
        self.student_figure = Figure(figsize=(10, 6))
        self.student_figure.patch.set_facecolor(colors.BG_CARD)
        self.student_canvas = FigureCanvas(self.student_figure)
        c_layout.addWidget(self.student_canvas)
        layout.addWidget(chart_card)

        return widget

    def generate_student_chart(self):
        self.student_figure.clear()
        ax = self.student_figure.add_subplot(111)
        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()

        class_id = self.class_combo.currentData()
        if not class_id:
            ax.text(0.5, 0.5, 'Veuillez sélectionner une classe', ha='center', va='center', fontsize=14, color=colors.TEXT_SECONDARY)
            ax.axis('off')
            self.student_canvas.draw()
            return

        db = DatabaseManager()
        grades = []
        max_score = 20
        active_year_label = "N/A"
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            active_year_id, active_year_label = repo.get_active_year_context()
            max_score = repo.get_class_max_score(class_id)

            if active_year_id:
                grades = repo.get_grades_for_class(class_id, active_year_id)

        if not grades:
            ax.text(0.5, 0.5, 'Aucune note disponible pour cette classe', ha='center', va='center', fontsize=14, color=colors.TEXT_SECONDARY)
            ax.axis('off')
        else:
            ax.set_facecolor(colors.BG_CARD)
            grades = [max(0, min(float(g), float(max_score))) for g in grades]
            step = max_score / 5
            bins = [round(i * step, 2) for i in range(6)]
            ax.hist(grades, bins=bins, color=colors.PRIMARY, edgecolor='white', alpha=0.8, rwidth=0.9)
            ax.set_xlim(0, max_score)
            ax.set_xlabel(f'Notes sur {int(max_score)} (Intervalle)', fontsize=10, fontweight='bold', color=colors.TEXT_SECONDARY)
            ax.set_ylabel("Nombre d'étudiants", fontsize=10, fontweight='bold', color=colors.TEXT_SECONDARY)
            ax.set_title(f'Distribution des Notes sur {int(max_score)} - {self.class_combo.currentText()}', fontsize=12, fontweight='bold', pad=20, color=colors.TEXT_PRIMARY)
            ax.text(0.5, 1.01, f"Année scolaire: {active_year_label}", transform=ax.transAxes, ha='center', va='bottom', fontsize=9, color=colors.TEXT_SECONDARY)
            ax.grid(axis='y', alpha=0.3, linestyle='--')

            avg = sum(grades) / len(grades)
            ax.axvline(avg, color=colors.DANGER, linestyle='--', linewidth=2, label=f'Moyenne Classe: {avg:.1f}/{int(max_score)}')
            legend = ax.legend()
            ax.tick_params(axis='x', colors=colors.TEXT_SECONDARY)
            ax.tick_params(axis='y', colors=colors.TEXT_SECONDARY)
            for spine in ax.spines.values(): spine.set_color(colors.BORDER)
            if legend:
                for text in legend.get_texts(): text.set_color(colors.TEXT_SECONDARY)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

        self.student_figure.tight_layout()
        self.student_canvas.draw()

    def create_excel_exports_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        grid = QGridLayout()
        grid.setSpacing(20)

        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()

        btn_financial = self.create_export_button("💰 Rapport Financier", "Export complet des recettes et dépenses", colors.SUCCESS)
        btn_financial.clicked.connect(lambda: self.export_excel("financial"))
        grid.addWidget(btn_financial, 0, 0)

        btn_students = self.create_export_button("👨‍🎓 Liste des Étudiants", "Export de tous les élèves inscrits", colors.PRIMARY)
        btn_students.clicked.connect(lambda: self.export_excel("students"))
        grid.addWidget(btn_students, 0, 1)

        btn_attendance = self.create_export_button("📅 Rapport d'Assiduité", "Statistiques de présence par classe", colors.WARNING)
        btn_attendance.clicked.connect(lambda: self.export_excel("attendance"))
        grid.addWidget(btn_attendance, 1, 0)

        secondary_color = colors.SECONDARY if hasattr(colors, "SECONDARY") else colors.PRIMARY
        btn_grades = self.create_export_button("📝 Relevé de Notes", "Tableau global des notes et moyennes", secondary_color)
        btn_grades.clicked.connect(lambda: self.export_excel("grades"))
        grid.addWidget(btn_grades, 1, 1)

        btn_comprehensive_pdf = self.create_export_button("📄 Rapport PDF Global", "Résumé financier et académique avec en-tête officiel", colors.PRIMARY)
        btn_comprehensive_pdf.clicked.connect(self.export_comprehensive_pdf)
        grid.addWidget(btn_comprehensive_pdf, 2, 0, 1, 2)

        layout.addLayout(grid)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(20)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{ border: none; border-radius: 4px; text-align: center; background-color: {colors.BG_MAIN}; color: {colors.TEXT_PRIMARY}; }}
            QProgressBar::chunk {{ background-color: {colors.PRIMARY}; border-radius: 4px; }}
        """)
        self.progress_bar.hide()
        layout.addWidget(self.progress_bar)

        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setStyleSheet(f"font-size: 13px; color: {colors.TEXT_SECONDARY}; font-weight: bold;")
        layout.addWidget(self.status_label)

        layout.addStretch()
        return widget

    def create_export_button(self, title, description, color):
        btn = QPushButton()
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setMinimumHeight(110)
        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()
        btn.setStyleSheet(f"""
            QPushButton {{ background-color: {colors.BG_CARD}; border: 1px solid {colors.BORDER}; border-radius: 12px; text-align: left; padding: 15px; border-left: 6px solid {color}; }}
            QPushButton:hover {{ background-color: {colors.BG_MAIN}; border: 1px solid {color}; border-left: 6px solid {color}; }}
        """)

        layout = QVBoxLayout(btn)
        layout.setContentsMargins(15, 10, 15, 10)

        title_label = QLabel(title)
        title_label.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {color}; border: none; background: transparent;")

        desc_label = QLabel(description)
        desc_label.setStyleSheet(f"font-size: 12px; color: {colors.TEXT_SECONDARY}; border: none; background: transparent;")

        layout.addWidget(title_label)
        layout.addWidget(desc_label)
        layout.addStretch()

        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15); shadow.setColor(QColor(0, 0, 0, 10)); shadow.setOffset(0, 4)
        btn.setGraphicsEffect(shadow)

        return btn

    def export_excel(self, report_type):
        default_name = f"Rapport_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
        params = {}

        if report_type == "financial":
            selected_period = self.period_combo.currentText() if hasattr(self, 'period_combo') else "12 derniers mois"
            safe_period = selected_period.encode('ascii', 'ignore').decode('ascii').replace(" ", "_").replace("'", "").replace("/", "-")
            if not safe_period: safe_period = "periode"
            default_name = f"Rapport_Financier_{safe_period}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
            params["period"] = selected_period
        elif report_type == "students":
            default_name = f"Liste_Etudiants_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
        elif report_type == "attendance":
            default_name = f"Rapport_Assiduite_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
        elif report_type == "grades":
            default_name = f"Rapport_Notes_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"

        output_path, _ = QFileDialog.getSaveFileName(self, "Enregistrer le rapport", default_name, "Excel Files (*.xlsx)")
        if not output_path: return

        params["output_path"] = output_path
        self.progress_bar.show()
        self.progress_bar.setValue(0)
        self.status_label.setText("⏳ Génération du rapport en cours...")

        AppLogger.info("AdvancedReports", f"Début de l'export Excel: {report_type}")

        if report_type in ["financial", "students", "attendance", "grades"]:
            self.worker = ReportWorker(f"excel_{report_type}", params)
        else:
            QMessageBox.information(self, "En développement", f"Export '{report_type}' sera disponible prochainement!")
            self.progress_bar.hide()
            self.status_label.setText("")
            return

        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.finished.connect(self.on_export_finished)
        self.worker.error.connect(self.on_export_error)
        self.worker.start()

    def export_comprehensive_pdf(self):
        selected_period = self.period_combo.currentText() if hasattr(self, 'period_combo') else "12 derniers mois"
        try:
            pdf = self._build_comprehensive_pdf(selected_period)
            mode = get_report_output_mode("advanced_comprehensive_pdf_mode", ADVANCED_COMPREHENSIVE_PDF_MODE)
            AppLogger.info("AdvancedReports", "Génération du PDF compréhensif terminée")

            output_pdf(
                pdf,
                self,
                default_name=f"Rapport_Complet_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.pdf",
                mode=mode,
                dialog_title="Enregistrer le rapport PDF",
                success_save_message="Rapport PDF généré avec succès.",
                success_print_message="Rapport PDF envoyé à l'imprimante.",
            )
        except Exception as e:
            AppLogger.error("AdvancedReports", "Erreur PDF compréhensif", e)
            QMessageBox.critical(self, "Erreur", f"Échec de la création du PDF: {e}")

    def on_export_finished(self, filepath):
        self.progress_bar.hide()
        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()
        self.status_label.setText(f"✅ Export réussi: {os.path.basename(filepath)}")
        self.status_label.setStyleSheet(f"color: {colors.SUCCESS}; font-weight: bold;")

        AppLogger.info("AdvancedReports", f"Export Excel terminé: {filepath}")

        reply = QMessageBox.question(
            self, "Export Réussi",
            f"Le fichier a été généré avec succès!\n\n{filepath}\n\nVoulez-vous l'ouvrir?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                if os.name == 'nt': os.startfile(filepath)
                else:
                    import subprocess
                    subprocess.call(('xdg-open', filepath))
            except Exception as e:
                AppLogger.warning("AdvancedReports", f"Impossible d'ouvrir le fichier exporté: {e}")

    def on_export_error(self, error):
        self.progress_bar.hide()
        colors = ThemeManager.get_colors() if THEME_AVAILABLE else Colors()
        self.status_label.setText("❌ Erreur lors de l'export")
        self.status_label.setStyleSheet(f"color: {colors.DANGER}; font-weight: bold;")
        QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export:\n{error}")

    def load_classes_combo(self):
        self.class_combo.clear()
        self.class_combo.addItem("-- Sélectionnez une classe --", None)
        try:
            db = DatabaseManager()
            with db.get_connection() as conn:
                for class_id, class_name in FinanceRepository(conn).list_classes():
                    self.class_combo.addItem(str(class_name or "-"), class_id)
        except Exception as e:
            AppLogger.error("AdvancedReports", "Erreur lors du chargement des classes", e)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AdvancedReportsWindow()
    window.show()
    sys.exit(app.exec())
