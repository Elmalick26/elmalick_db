"""
analytics_dashboard.py — Phase 6.1+ (consolidé avec advanced_reports)
لوحة تحليلات وتقارير متكاملة.

• KPI: moyenne générale, assiduité, recouvrement, élèves en difficulté
• Tab 1 — Notes par matière     (bar chart normalisé /20)
• Tab 2 — Distribution des notes par classe  (histogramme)
• Tab 3 — Évolution de la présence mensuelle (line chart)
• Tab 4 — Tableau financier: recouvrement (pie) + évolution mensuelle (bar)
• Tab 5 — Exports Excel (finances, étudiants, assiduité, notes) + PDF global
"""

from __future__ import annotations

import os
import sys
from datetime import datetime

import matplotlib
import matplotlib.pyplot as plt  # noqa: E402
from fpdf import FPDF
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas  # noqa: E402
from matplotlib.figure import Figure  # noqa: E402
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QFrame,
    QGraphicsDropShadowEffect,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSplitter,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from app_logger import AppLogger  # noqa: E402
from database_setup import DatabaseManager  # noqa: E402
from pdf_report_style import (  # noqa: E402
    apply_grades_sheet_header,
    apply_table_body_style,
    apply_table_header_style,
    get_school_info_row,
    set_zebra_row_fill,
)
from print_export_service import get_report_output_mode, output_pdf  # noqa: E402
from repositories.analytics_repo import AnalyticsRepository  # noqa: E402
from repositories.finance_repo import FinanceRepository  # noqa: E402
from services.grade_service import GradeService  # noqa: E402
from ui_styles import Colors, KpiCard, ModuleHeaderWidget, ThemeManager, get_tabs_style  # noqa: E402

COMPREHENSIVE_PDF_MODE = get_report_output_mode("advanced_comprehensive_pdf_mode", "save")

# ── Chart palette — dark theme for matplotlib (independent of UI theme) ────
# These constants are ONLY used in matplotlib drawing code (chart backgrounds/axes).
# All Qt widget styling uses ThemeManager instead.
_CHART_BG = "#161b2e"
_CHART_CARD = "#1e2433"
_CHART_CARD2 = "#252b3b"
_CHART_TEXT = "#e0e0e0"
_CHART_MUTED = "#9ca3af"
_CHART_BORDER = "#374151"
_CHART_ACCENT = "#4fc3f7"


def _drop_shadow(widget) -> None:
    """Attach a subtle drop-shadow to *widget*."""
    eff = QGraphicsDropShadowEffect()
    eff.setBlurRadius(12)
    eff.setColor(QColor(0, 0, 0, 55))
    eff.setOffset(0, 3)
    widget.setGraphicsEffect(eff)


# ──────────────────────────────────────────────────────────────────────────
# Worker: Excel report generation (runs in a background thread)
# ──────────────────────────────────────────────────────────────────────────
class ReportWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, report_type: str, params: dict):
        super().__init__()
        self.report_type = report_type
        self.params = params

    # -- helpers -----------------------------------------------------------

    @staticmethod
    def _name_expr(cols: set, alias: str = "s") -> str:
        parts: list[str] = []
        if {"first_name_fr", "last_name_fr"}.issubset(cols):
            parts.append(f"TRIM(NULLIF({alias}.first_name_fr,'') || ' ' || NULLIF({alias}.last_name_fr,''))")
        if "student_name" in cols:
            parts.append(f"NULLIF({alias}.student_name,'')")
        if "first_name_fr" in cols:
            parts.append(f"NULLIF({alias}.first_name_fr,'')")
        if "last_name_fr" in cols:
            parts.append(f"NULLIF({alias}.last_name_fr,'')")
        return f"COALESCE({', '.join(parts)}, 'N/A')" if parts else "'N/A'"

    @staticmethod
    def _hdr_fill() -> PatternFill:
        return PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    @staticmethod
    def _hdr_font() -> Font:
        return Font(color="FFFFFF", bold=True)

    def _save(self, wb: Workbook, name: str) -> str:
        fp = self.params.get("output_path") or os.path.join(os.getcwd(), name)
        wb.save(fp)
        return fp

    # -- run dispatcher ----------------------------------------------------

    def run(self) -> None:
        try:
            dispatch = {
                "excel_financial": self._gen_financial,
                "excel_students": self._gen_students,
                "excel_attendance": self._gen_attendance,
                "excel_grades": self._gen_grades,
            }
            fn = dispatch.get(self.report_type)
            if fn:
                self.finished.emit(fn())
            else:
                self.error.emit(f"Type de rapport inconnu: {self.report_type}")
        except Exception as exc:
            AppLogger.error("ReportWorker", f"{self.report_type}: {exc}")
            self.error.emit(str(exc))

    # -- generators --------------------------------------------------------

    def _gen_financial(self) -> str:
        period = self.params.get("period", "12 derniers mois")
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport Financier"
        hf, hfont = self._hdr_fill(), self._hdr_font()

        ws["A1"] = "RAPPORT FINANCIER COMPLET"
        ws["A1"].font = Font(bold=True, size=16, color="1E293B")
        ws.merge_cells("A1:F1")
        ws["A2"] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells("A2:F2")
        ws["A3"] = f"Période: {period}"
        ws.merge_cells("A3:F3")
        self.progress.emit(10)

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            income_rows = repo.get_monthly_income_totals(period)
            expense_rows = repo.get_monthly_expense_totals(period)

        # Recettes
        ws["A4"] = "RECETTES"
        ws["A4"].font = Font(bold=True, size=12)
        for col, hdr in enumerate(["Mois", "Nb Paiements", "Montant Total"], 1):
            c = ws.cell(row=5, column=col, value=hdr)
            c.fill = hf
            c.font = hfont
        row, total_inc = 6, 0
        for month, cnt, total in income_rows:
            ws[f"A{row}"] = month
            ws[f"B{row}"] = cnt
            ws[f"C{row}"] = f"{total or 0:,.0f} FCFA"
            total_inc += total or 0
            row += 1
        ws[f"A{row}"] = "TOTAL RECETTES"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"] = f"{total_inc:,.0f} FCFA"
        ws[f"C{row}"].font = Font(bold=True, color="10B981")
        self.progress.emit(50)

        # Dépenses
        r0 = row + 2
        ws[f"A{r0}"] = "DÉPENSES"
        ws[f"A{r0}"].font = Font(bold=True, size=12)
        for col, hdr in enumerate(["Mois", "Nb Dépenses", "Montant Total"], 1):
            c = ws.cell(row=r0 + 1, column=col, value=hdr)
            c.fill = hf
            c.font = hfont
        r, total_exp = r0 + 2, 0
        for month, cnt, total in expense_rows:
            ws[f"A{r}"] = month
            ws[f"B{r}"] = cnt
            ws[f"C{r}"] = f"{total or 0:,.0f} FCFA"
            total_exp += total or 0
            r += 1
        ws[f"A{r}"] = "TOTAL DÉPENSES"
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"C{r}"] = f"{total_exp:,.0f} FCFA"
        ws[f"C{r}"].font = Font(bold=True, color="EF4444")
        bal = total_inc - total_exp
        ws[f"A{r + 2}"] = "SOLDE NET"
        ws[f"A{r + 2}"].font = Font(bold=True, size=13)
        ws[f"C{r + 2}"] = f"{bal:,.0f} FCFA"
        ws[f"C{r + 2}"].font = Font(bold=True, size=13, color="10B981" if bal >= 0 else "EF4444")
        for col, w in zip("ABC", [15, 15, 22]):
            ws.column_dimensions[col].width = w
        self.progress.emit(100)
        safe = period.encode("ascii", "ignore").decode("ascii").replace(" ", "_").replace("'", "") or "periode"
        return self._save(wb, f"Rapport_Financier_{safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    def _gen_students(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Étudiants"
        hf, hfont = self._hdr_fill(), self._hdr_font()

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            year_id, year_label = repo.get_active_year_context()
            cols = repo.get_student_columns()
            name_expr = self._name_expr(cols)
            bp = "s.birth_place" if "birth_place" in cols else "''"
            pr = "s.parent_name" if "parent_name" in cols else "''"
            ph = "s.phone" if "phone" in cols else "''"
            ad = "s.address" if "address" in cols else "''"
            en = "s.registration_date" if "registration_date" in cols else "''"

        ws["A1"] = "LISTE DES ÉTUDIANTS"
        ws["A1"].font = Font(bold=True, size=16, color="1E293B")
        ws.merge_cells("A1:L1")
        ws["A2"] = f"Année scolaire: {year_label}"
        ws.merge_cells("A2:L2")
        hdrs = [
            "ID",
            "N°",
            "Nom Complet",
            "Classe",
            "Sexe",
            "Naissance",
            "Lieu",
            "Parent",
            "Tél.",
            "Adresse",
            "Inscription",
            "Statut",
        ]
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.fill = hf
            c.font = hfont
        self.progress.emit(25)

        with db.get_connection() as conn:
            cur = conn.cursor()
            if year_id:
                cur.execute(
                    f"""SELECT s.id,COALESCE(CAST(scn.class_number AS VARCHAR),''),
                    {name_expr} AS n,c.class_name_fr,
                    CASE WHEN s.gender IN ('0','M') THEN 'M' ELSE 'F' END,
                    CAST(s.birth_date AS VARCHAR),{bp},{pr},{ph},{ad},
                    CAST({en} AS VARCHAR),s.status
                    FROM Students s
                    JOIN StudentClassNumbers scn ON scn.student_id=s.id AND scn.year_id=%s
                    JOIN Classes c ON scn.class_id=c.id
                    WHERE s.status='Active'
                    ORDER BY c.sort_order,COALESCE(scn.class_number,9999),n""",
                    (year_id,),
                )
            else:
                cur.execute(
                    f"""SELECT s.id,'',{name_expr} AS n,'N/A',
                    CASE WHEN s.gender IN ('0','M') THEN 'M' ELSE 'F' END,
                    CAST(s.birth_date AS VARCHAR),{bp},{pr},{ph},{ad},
                    CAST({en} AS VARCHAR),s.status
                    FROM Students s WHERE s.status='Active' ORDER BY n"""
                )
            rows = cur.fetchall()

        for i, student in enumerate(rows, 5):
            for col, val in enumerate(student, 1):
                ws.cell(row=i, column=col, value=val)
        for col, w in zip("ABCDEFGHIJKL", [8, 10, 28, 18, 8, 14, 18, 22, 14, 24, 14, 10]):
            ws.column_dimensions[col].width = w
        self.progress.emit(100)
        return self._save(wb, f"Liste_Etudiants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    def _gen_attendance(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Assiduité"
        hf, hfont = self._hdr_fill(), self._hdr_font()
        ws["A1"] = "RAPPORT D'ASSIDUITÉ"
        ws["A1"].font = Font(bold=True, size=16, color="1E293B")
        ws.merge_cells("A1:G1")
        ws["A2"] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells("A2:G2")
        for col, h in enumerate(["Classe", "Étudiants", "Présents", "Absents", "Retards", "Total", "Taux Abs%"], 1):
            c = ws.cell(row=4, column=col, value=h)
            c.fill = hf
            c.font = hfont
        self.progress.emit(20)

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            year_id, year_label = repo.get_active_year_context()
            ws["A3"] = f"Année: {year_label}"
            ws.merge_cells("A3:G3")
            data = repo.get_attendance_summary_by_class(year_id) if year_id else []

        for i, row in enumerate(data, 5):
            cls, cnt, pres, abs_, late, total = row
            rate = (abs_ or 0) / max(total or 1, 1) * 100
            for col, val in enumerate(
                [cls or "-", cnt or 0, pres or 0, abs_ or 0, late or 0, total or 0, round(rate, 2)],
                1,
            ):
                ws.cell(row=i, column=col, value=val)

        for col, w in zip("ABCDEFG", [24, 12, 12, 12, 12, 12, 16]):
            ws.column_dimensions[col].width = w
        self.progress.emit(100)
        return self._save(wb, f"Rapport_Assiduite_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    def _gen_grades(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Notes"
        hf, hfont = self._hdr_fill(), self._hdr_font()
        ws["A1"] = "RAPPORT GLOBAL DES NOTES"
        ws["A1"].font = Font(bold=True, size=16, color="1E293B")
        ws.merge_cells("A1:F1")
        ws["A2"] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells("A2:F2")
        for col, h in enumerate(["Classe", "Étudiants", "Nb Notes", "Min", "Moyenne", "Max"], 1):
            c = ws.cell(row=4, column=col, value=h)
            c.fill = hf
            c.font = hfont
        self.progress.emit(20)

        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            year_id, year_label = repo.get_active_year_context()
            ws["A3"] = f"Année: {year_label}"
            ws.merge_cells("A3:F3")
            data = repo.get_grades_summary_by_class(year_id) if year_id else []

        for i, row in enumerate(data, 5):
            cls, cnt, nb, mn, avg, mx = row
            for col, val in enumerate(
                [
                    cls or "-",
                    cnt or 0,
                    nb or 0,
                    round(mn, 2) if mn is not None else "-",
                    round(avg, 2) if avg is not None else "-",
                    round(mx, 2) if mx is not None else "-",
                ],
                1,
            ):
                ws.cell(row=i, column=col, value=val)

        for col, w in zip("ABCDEF", [24, 12, 12, 10, 12, 10]):
            ws.column_dimensions[col].width = w
        self.progress.emit(100)
        return self._save(wb, f"Rapport_Notes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


# ──────────────────────────────────────────────────────────────────────────
# Worker: KPI + chart data (runs in background thread)
# ──────────────────────────────────────────────────────────────────────────
class AnalyticsWorker(QThread):
    grades_ready = pyqtSignal(list, list, list)  # (names, normalized_avgs, coefficients)
    attendance_ready = pyqtSignal(dict)  # {"YYYY-MM": rate_pct}
    finance_ready = pyqtSignal(float, float)  # (total_paid, total_due)
    error_signal = pyqtSignal(str)

    def __init__(self, year_id: int, class_id: int | None = None):
        super().__init__()
        self.year_id = year_id
        self.class_id = class_id

    def run(self) -> None:
        try:
            db = DatabaseManager()
            with db.get_connection() as conn:
                repo = AnalyticsRepository(conn)
                self._load_grades(repo)
                self._load_attendance(repo)
                self._load_finance(repo)
        except Exception as exc:
            AppLogger.error("AnalyticsDashboard", f"Worker error: {exc}")
            self.error_signal.emit(str(exc))

    def _load_grades(self, repo: AnalyticsRepository) -> None:
        rows = repo.get_grades_by_subject(self.year_id, self.class_id)
        names = [r[0] for r in rows]
        averages = [float(r[1]) if r[1] is not None else 0.0 for r in rows]
        max_scores = [float(r[3]) for r in rows]
        normalized = [avg / mx * 20 if mx else 0 for avg, mx in zip(averages, max_scores)]
        self.grades_ready.emit(names, normalized, [float(r[2]) for r in rows])

    def _load_attendance(self, repo: AnalyticsRepository) -> None:
        monthly: dict[str, float] = {}
        for month, total, present in repo.get_monthly_attendance_rate(self.year_id, self.class_id):
            if total and total > 0:
                monthly[month] = round(present / total * 100, 1)
        self.attendance_ready.emit(monthly)

    def _load_finance(self, repo: AnalyticsRepository) -> None:
        paid, due = repo.get_finance_summary(self.year_id, self.class_id)
        self.finance_ready.emit(paid, due)


# ──────────────────────────────────────────────────────────────────────────
# Chart widgets — theme-aware palette
# ──────────────────────────────────────────────────────────────────────────
def _chart_colors():
    c = ThemeManager.get_colors()
    return {
        "figure": c.BG_CARD,
        "panel": c.BG_CARD,
        "panel_alt": c.BG_MAIN,
        "text": c.TEXT_PRIMARY,
        "muted": c.TEXT_SECONDARY,
        "border": c.BORDER,
        "accent": c.PRIMARY,
        "danger": c.DANGER,
    }


def _theme_figure(figsize=(6, 4)) -> Figure:
    return Figure(figsize=figsize, tight_layout=True, facecolor=_chart_colors()["figure"])


class GradesBarChart(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.fig = _theme_figure((7, 4))
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvas(self.fig)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self.canvas)
        self._draw_empty()

    def _draw_empty(self) -> None:
        c = _chart_colors()
        self.ax.clear()
        self.ax.set_facecolor(c["panel"])
        self.ax.text(
            0.5,
            0.5,
            "En attente de données…",
            ha="center",
            va="center",
            color=c["muted"],
            fontsize=12,
            transform=self.ax.transAxes,
        )
        self.ax.axis("off")
        self.canvas.draw()

    def update_data(self, names: list, averages: list, coefficients: list) -> None:
        c = _chart_colors()
        self.ax.clear()
        if not names:
            self._draw_empty()
            return
        colors = plt.cm.RdYlGn([a / 20 for a in averages])  # type: ignore[attr-defined]
        bars = self.ax.barh(names, averages, color=colors, edgecolor=c["border"])
        self.ax.set_xlim(0, 20)
        self.ax.set_xlabel("Moyenne /20", color=c["muted"])
        self.ax.set_title("Moyennes par Matière", color=c["text"], fontsize=13, pad=10)
        self.ax.tick_params(colors=c["muted"])
        self.ax.set_facecolor(c["panel"])
        self.fig.set_facecolor(c["figure"])
        for spine in self.ax.spines.values():
            spine.set_edgecolor(c["border"])
        for bar, avg in zip(bars, averages):
            self.ax.text(
                bar.get_width() + 0.2,
                bar.get_y() + bar.get_height() / 2,
                f"{avg:.1f}",
                va="center",
                color=c["text"],
                fontsize=9,
            )
        self.canvas.draw()


class AttendanceLineChart(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.fig = _theme_figure((7, 3.5))
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvas(self.fig)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self.canvas)
        self._draw_empty()

    def _draw_empty(self) -> None:
        c = _chart_colors()
        self.ax.clear()
        self.ax.set_facecolor(c["panel"])
        self.ax.text(
            0.5,
            0.5,
            "Chargement…",
            ha="center",
            va="center",
            color=c["muted"],
            fontsize=12,
            transform=self.ax.transAxes,
        )
        self.ax.axis("off")
        self.canvas.draw()

    def update_data(self, monthly: dict) -> None:
        c = _chart_colors()
        self.ax.clear()
        if not monthly:
            self.ax.text(
                0.5,
                0.5,
                "Aucune donnée de présence",
                ha="center",
                va="center",
                color=c["muted"],
                fontsize=11,
                transform=self.ax.transAxes,
            )
            self.ax.axis("off")
            self.canvas.draw()
            return
        months = list(monthly.keys())
        rates = list(monthly.values())
        labels = [m[5:] for m in months]
        self.ax.plot(labels, rates, marker="o", linewidth=2, color=c["accent"], markersize=6)
        self.ax.fill_between(range(len(rates)), rates, alpha=0.15, color=c["accent"])
        self.ax.axhline(80, linestyle="--", color=c["danger"], linewidth=1, label="Seuil 80%")
        self.ax.set_ylim(0, 105)
        self.ax.set_ylabel("Taux de présence (%)", color=c["muted"])
        self.ax.set_title("Évolution de la Présence Mensuelle", color=c["text"], fontsize=13, pad=10)
        self.ax.tick_params(colors=c["muted"])
        self.ax.set_facecolor(c["panel"])
        self.fig.set_facecolor(c["figure"])
        for spine in self.ax.spines.values():
            spine.set_edgecolor(c["border"])
        self.ax.legend(facecolor=c["panel"], labelcolor=c["muted"])
        self.canvas.draw()


class FinancePieChart(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.fig = _theme_figure((5, 4))
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvas(self.fig)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self.canvas)
        self._draw_empty()

    def _draw_empty(self) -> None:
        c = _chart_colors()
        self.ax.clear()
        self.ax.set_facecolor(c["panel"])
        self.ax.text(
            0.5,
            0.5,
            "Chargement…",
            ha="center",
            va="center",
            color=c["muted"],
            fontsize=12,
            transform=self.ax.transAxes,
        )
        self.ax.axis("off")
        self.canvas.draw()

    def update_data(self, paid: float, total_due: float) -> None:
        c = _chart_colors()
        self.ax.clear()
        remaining = max(0.0, total_due - paid)
        if total_due <= 0:
            self.ax.text(
                0.5,
                0.5,
                "Aucune donnée financière",
                ha="center",
                va="center",
                color=c["muted"],
                fontsize=11,
                transform=self.ax.transAxes,
            )
            self.ax.axis("off")
            self.canvas.draw()
            return
        _, _, autotexts = self.ax.pie(
            [paid, remaining],
            labels=[f"Payé\n{paid:,.0f} FCFA", f"Restant\n{remaining:,.0f} FCFA"],
            autopct="%1.1f%%",
            colors=["#66bb6a", "#ef5350"],
            explode=(0.05, 0),
            startangle=90,
            textprops={"color": c["text"], "fontsize": 9},
        )
        for at in autotexts:
            at.set_fontsize(10)
            at.set_color(c["text"])
        self.ax.set_title("Recouvrement des Frais", color=c["text"], fontsize=13, pad=10)
        self.fig.set_facecolor(c["figure"])
        self.canvas.draw()


# KpiCard imported from ui_styles — see shared component at top of file


# ──────────────────────────────────────────────────────────────────────────
# Main Window
# ──────────────────────────────────────────────────────────────────────────
class AnalyticsDashboardWindow(QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Analytics & Rapports — Tableau de Bord")
        ThemeManager.apply_theme(self)
        self._year_id: int | None = None
        self._class_id: int | None = None
        self._worker: AnalyticsWorker | None = None
        self._report_worker: ReportWorker | None = None
        self._svc = GradeService()
        self._setup_ui()
        self._load_filters()

    def _setup_ui(self) -> None:
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(16, 12, 16, 12)
        root.setSpacing(10)
        # ThemeManager applies via apply_theme() in __init__ — no hardcoded background needed
        root.addWidget(self._build_header())
        root.addLayout(self._build_kpi_row())
        root.addWidget(self._build_tabs())

    def _build_header(self) -> ModuleHeaderWidget:
        hdr = ModuleHeaderWidget(
            icon="📊",
            title="ANALYTICS & RAPPORTS",
            subtitle="Statistiques · Graphiques · Export Excel & PDF",
        )
        hl = hdr.layout()
        # Append year/class filters and refresh button
        c = ThemeManager.get_colors()
        _combo_style = (
            f"QComboBox {{ background:{c.INPUT_BG}; color:{c.TEXT_PRIMARY}; border:1px solid {c.BORDER}; "
            "border-radius:4px; padding:4px 8px; min-width:120px; } "
            "QComboBox::drop-down { border:none; } "
            f"QComboBox QAbstractItemView {{ background:{c.BG_CARD}; color:{c.TEXT_PRIMARY}; }}"
        )
        lbl_yr = QLabel("Année:")
        lbl_yr.setStyleSheet(f"color:{c.HEADER_TEXT}; background:transparent;")
        hl.addWidget(lbl_yr)
        self.cmb_year = QComboBox()
        self.cmb_year.setMinimumWidth(140)
        self.cmb_year.setStyleSheet(_combo_style)
        self.cmb_year.currentIndexChanged.connect(self._on_year_changed)
        hl.addWidget(self.cmb_year)
        lbl_cl = QLabel("Classe:")
        lbl_cl.setStyleSheet(f"color:{c.HEADER_TEXT}; background:transparent;")
        hl.addWidget(lbl_cl)
        self.cmb_class = QComboBox()
        self.cmb_class.setMinimumWidth(130)
        self.cmb_class.setStyleSheet(_combo_style)
        self.cmb_class.addItem("Toutes les classes", None)
        self.cmb_class.currentIndexChanged.connect(self._refresh)
        hl.addWidget(self.cmb_class)
        btn = QPushButton("🔄 Actualiser")
        btn.setStyleSheet(
            "QPushButton { background:#2962ff; color:white; border-radius:6px; "
            "padding:6px 14px; font-weight:bold; border:none; } "
            "QPushButton:hover { background:#1565c0; }"
        )
        btn.clicked.connect(self._refresh)
        hl.addWidget(btn)
        return hdr

    def _build_kpi_row(self) -> QHBoxLayout:
        row = QHBoxLayout()
        self.kpi_avg = KpiCard("🎓", "Moyenne Générale", "—", "#66bb6a", theme_mode="default")
        self.kpi_att = KpiCard("📅", "Taux de Présence", "—", _CHART_ACCENT, theme_mode="default")
        self.kpi_paid = KpiCard("💰", "Taux de Recouvrement", "—", "#ffa726", theme_mode="default")
        self.kpi_risk = KpiCard("⚠️", "Elèves en Difficulté", "—", "#ef5350", theme_mode="default")
        for card in (self.kpi_avg, self.kpi_att, self.kpi_paid, self.kpi_risk):
            row.addWidget(card)
        return row

    def _build_tabs(self) -> QTabWidget:
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(get_tabs_style())
        self.tabs.addTab(self._tab_notes(), "📚 Notes par Matière")
        self.tabs.addTab(self._tab_histogram(), "📈 Distribution Notes")
        self.tabs.addTab(self._tab_attendance(), "📅 Présences Mensuelles")
        self.tabs.addTab(self._tab_finance(), "💰 Tableau Financier")
        self.tabs.addTab(self._tab_exports(), "📑 Exports Excel & PDF")
        self.tabs.currentChanged.connect(self._on_tab_changed)
        return self.tabs

    def _tab_notes(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(6, 6, 6, 6)
        self.grades_chart = GradesBarChart()
        lay.addWidget(self.grades_chart)
        return w

    def _tab_histogram(self) -> QWidget:
        w = QWidget()
        c = ThemeManager.get_colors()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(10)
        ctrl = QFrame()
        ctrl.setStyleSheet(f"QFrame {{ background:{c.BG_CARD}; border-radius:8px; border:1px solid {c.BORDER}; }}")
        cl = QHBoxLayout(ctrl)
        cl.setContentsMargins(12, 8, 12, 8)
        lbl_c = QLabel("Classe:")
        self.cmb_hist = QComboBox()
        self.cmb_hist.setMinimumWidth(200)
        btn_h = QPushButton("📈 Analyser")
        btn_h.clicked.connect(self._draw_histogram)
        cl.addWidget(lbl_c)
        cl.addWidget(self.cmb_hist)
        cl.addWidget(btn_h)
        cl.addStretch()
        lay.addWidget(ctrl)
        self.hist_fig = Figure(figsize=(9, 5), tight_layout=True, facecolor=c.BG_CARD)
        self.hist_canvas = FigureCanvas(self.hist_fig)
        lay.addWidget(self.hist_canvas)
        return w

    def _tab_attendance(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(6, 6, 6, 6)
        self.att_chart = AttendanceLineChart()
        lay.addWidget(self.att_chart)
        return w

    def _tab_finance(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(10)
        c = ThemeManager.get_colors()
        ctrl = QFrame()
        ctrl.setStyleSheet(f"QFrame {{ background:{c.BG_CARD}; border-radius:8px; border:1px solid {c.BORDER}; }}")
        cl = QHBoxLayout(ctrl)
        cl.setContentsMargins(12, 8, 12, 8)
        lbl_p = QLabel("Période:")
        lbl_p.setStyleSheet("")
        self.cmb_period = QComboBox()
        self.cmb_period.addItems(["6 derniers mois", "12 derniers mois", "Année en cours"])
        self.cmb_period.setCurrentIndex(1)
        self.cmb_period.setMinimumWidth(160)
        btn_f = QPushButton("🔄 Actualiser")
        btn_f.clicked.connect(self._draw_finance_bar)
        cl.addWidget(lbl_p)
        cl.addWidget(self.cmb_period)
        cl.addWidget(btn_f)
        cl.addStretch()
        lay.addWidget(ctrl)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        pie_w = QWidget()
        pie_w.setStyleSheet(f"background: {c.BG_CARD}; border: 1px solid {c.BORDER}; border-radius: 10px;")
        pie_l = QVBoxLayout(pie_w)
        pie_l.setContentsMargins(0, 0, 4, 0)
        lbl_pie = QLabel("Recouvrement des Frais Scolaires")
        lbl_pie.setStyleSheet("font-weight:bold; font-size:12px;")
        lbl_pie.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pie_l.addWidget(lbl_pie)
        self.fin_chart = FinancePieChart()
        pie_l.addWidget(self.fin_chart)
        splitter.addWidget(pie_w)
        bar_w = QWidget()
        bar_w.setStyleSheet(f"background: {c.BG_CARD}; border: 1px solid {c.BORDER}; border-radius: 10px;")
        bar_l = QVBoxLayout(bar_w)
        bar_l.setContentsMargins(4, 0, 0, 0)
        lbl_bar = QLabel("Évolution Financière Mensuelle")
        lbl_bar.setStyleSheet("font-weight:bold; font-size:12px;")
        lbl_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        bar_l.addWidget(lbl_bar)
        self.fin_bar_fig = Figure(figsize=(7, 5), tight_layout=True, facecolor=c.BG_CARD)
        self.fin_bar_canvas = FigureCanvas(self.fin_bar_fig)
        bar_l.addWidget(self.fin_bar_canvas)
        splitter.addWidget(bar_w)
        splitter.setSizes([350, 550])
        lay.addWidget(splitter)
        return w

    def _tab_exports(self) -> QWidget:
        w = QWidget()
        c = ThemeManager.get_colors()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(16)
        grid = QGridLayout()
        grid.setSpacing(14)

        def _btn(title: str, desc: str, accent: str) -> QPushButton:
            b = QPushButton()
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.setMinimumHeight(90)
            b.setStyleSheet(
                f"QPushButton {{ background:{c.BG_CARD}; border:1px solid {c.BORDER}; border-radius:10px; "
                f"text-align:left; padding:10px; border-left:5px solid {accent}; }} "
                f"QPushButton:hover {{ background:{c.BG_MAIN}; border-left:5px solid {accent}; }}"
            )
            inner = QVBoxLayout(b)
            inner.setContentsMargins(12, 6, 12, 6)
            t = QLabel(title)
            t.setStyleSheet(f"font-size:14px; font-weight:bold; color:{accent}; border:none; background:transparent;")
            d = QLabel(desc)
            d.setStyleSheet(f"font-size:11px; color:{c.TEXT_SECONDARY}; border:none; background:transparent;")
            inner.addWidget(t)
            inner.addWidget(d)
            inner.addStretch()
            sh = QGraphicsDropShadowEffect()
            sh.setBlurRadius(10)
            sh.setColor(QColor(0, 0, 0, 40))
            sh.setOffset(0, 3)
            b.setGraphicsEffect(sh)
            return b

        b_fin = _btn("💰 Rapport Financier Excel", "Recettes et dépenses mensuelles", "#10b981")
        b_fin.clicked.connect(lambda: self._export_excel("financial"))
        grid.addWidget(b_fin, 0, 0)
        b_stu = _btn("👨‍🎓 Liste des Étudiants Excel", "Tous les élèves inscrits avec détails", "#3b82f6")
        b_stu.clicked.connect(lambda: self._export_excel("students"))
        grid.addWidget(b_stu, 0, 1)
        b_att = _btn("📅 Rapport d\'Assiduité Excel", "Statistiques de présence par classe", "#f59e0b")
        b_att.clicked.connect(lambda: self._export_excel("attendance"))
        grid.addWidget(b_att, 1, 0)
        b_gr = _btn("📝 Relevé de Notes Excel", "Tableau global des notes et moyennes", "#8b5cf6")
        b_gr.clicked.connect(lambda: self._export_excel("grades"))
        grid.addWidget(b_gr, 1, 1)
        b_pdf = _btn("📄 Rapport PDF Global", "Résumé financier & académique avec en-tête officiel", "#0ea5e9")
        b_pdf.clicked.connect(self._export_comprehensive_pdf)
        grid.addWidget(b_pdf, 2, 0, 1, 2)
        lay.addLayout(grid)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(16)
        self.progress_bar.setStyleSheet(
            "QProgressBar { border:none; border-radius:4px; } "
            "QProgressBar::chunk { background:#3b82f6; border-radius:4px; }"
        )
        self.progress_bar.hide()
        lay.addWidget(self.progress_bar)
        self.status_lbl = QLabel("")
        self.status_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_lbl.setStyleSheet("font-size:12px; font-weight:bold;")
        lay.addWidget(self.status_lbl)
        lay.addStretch()
        return w

    def _load_filters(self) -> None:
        try:
            db = DatabaseManager()
            with db.get_connection() as conn:
                repo = AnalyticsRepository(conn)
                years = repo.get_academic_years()
                self.cmb_year.blockSignals(True)
                self.cmb_year.clear()
                active_idx = 0
                for i, (yid, label, is_active) in enumerate(years):
                    self.cmb_year.addItem(label, yid)
                    if is_active:
                        active_idx = i
                self.cmb_year.setCurrentIndex(active_idx)
                self.cmb_year.blockSignals(False)
                if years:
                    self._year_id = years[active_idx][0]
                    self._load_classes(conn, self._year_id)
                self.cmb_hist.clear()
                self.cmb_hist.addItem("-- Sélectionnez une classe --", None)
                for cid, cname in FinanceRepository(conn).list_classes():
                    self.cmb_hist.addItem(str(cname or "-"), cid)
        except Exception as exc:
            AppLogger.error("AnalyticsDashboard", f"_load_filters: {exc}")

    def _load_classes(self, conn, year_id: int) -> None:
        self.cmb_class.blockSignals(True)
        self.cmb_class.clear()
        self.cmb_class.addItem("Toutes les classes", None)
        try:
            for cid, name in AnalyticsRepository(conn).get_classes_for_year(year_id):
                self.cmb_class.addItem(name, cid)
        except Exception as exc:
            AppLogger.error("AnalyticsDashboard", f"_load_classes: {exc}")
        finally:
            self.cmb_class.blockSignals(False)

    def _on_year_changed(self) -> None:
        self._year_id = self.cmb_year.currentData()
        if self._year_id:
            try:
                db = DatabaseManager()
                with db.get_connection() as conn:
                    self._load_classes(conn, self._year_id)
            except Exception as exc:
                AppLogger.error("AnalyticsDashboard", f"_on_year_changed: {exc}")
        self._refresh()

    def _on_tab_changed(self, idx: int) -> None:
        if idx == 3:  # Finance tab -> auto-refresh bar chart
            self._draw_finance_bar()

    def _refresh(self) -> None:
        self._year_id = self.cmb_year.currentData()
        self._class_id = self.cmb_class.currentData()
        if not self._year_id:
            return
        if self._worker and self._worker.isRunning():
            self._worker.quit()
        self._worker = AnalyticsWorker(self._year_id, self._class_id)
        self._worker.grades_ready.connect(self._on_grades)
        self._worker.attendance_ready.connect(self._on_attendance)
        self._worker.finance_ready.connect(self._on_finance)
        self._worker.error_signal.connect(lambda e: AppLogger.error("Analytics", e))
        self._worker.start()

    def _on_grades(self, names: list, averages: list, coefficients: list) -> None:
        self.grades_chart.update_data(names, averages, coefficients)
        if averages:
            gen_avg = self._svc.calculate_period_average(list(zip(averages, coefficients)))
            self.kpi_avg.set_value(f"{gen_avg:.2f}/20\n{self._svc.get_honor_mention(gen_avg)}")
            self.kpi_risk.set_value(str(sum(a < 10 for a in averages)))

    def _on_attendance(self, monthly: dict) -> None:
        self.att_chart.update_data(monthly)
        if monthly:
            self.kpi_att.set_value(f"{sum(monthly.values()) / len(monthly):.1f}%")

    def _on_finance(self, paid: float, total_due: float) -> None:
        self.fin_chart.update_data(paid, total_due)
        self.kpi_paid.set_value(f"{paid / total_due * 100:.1f}%" if total_due > 0 else "—")

    def _draw_histogram(self) -> None:
        c = _chart_colors()
        class_id = self.cmb_hist.currentData()
        self.hist_fig.patch.set_facecolor(c["figure"])
        self.hist_fig.clear()
        ax = self.hist_fig.add_subplot(111)
        ax.set_facecolor(c["panel"])
        if not class_id:
            ax.text(0.5, 0.5, "Sélectionnez une classe", ha="center", va="center", fontsize=13, color=c["muted"])
            ax.axis("off")
            self.hist_canvas.draw()
            return
        grades, max_score, year_label = [], 20, "—"
        try:
            db = DatabaseManager()
            with db.get_connection() as conn:
                repo = AnalyticsRepository(conn)
                year_id, year_label = repo.get_active_year_context()
                max_score = repo.get_class_max_score(class_id) or 20
                if year_id:
                    grades = repo.get_grades_for_class(class_id, year_id)
        except Exception as exc:
            AppLogger.error("AnalyticsDashboard", f"histogram: {exc}")
        if not grades:
            ax.text(0.5, 0.5, "Aucune note disponible", ha="center", va="center", fontsize=13, color=c["muted"])
            ax.axis("off")
            self.hist_canvas.draw()
            return
        grades = [max(0.0, min(float(g), float(max_score))) for g in grades]
        step = max_score / 5
        bins = [round(i * step, 2) for i in range(6)]
        ax.hist(grades, bins=bins, color=c["accent"], edgecolor=c["border"], alpha=0.85, rwidth=0.88)
        ax.set_xlim(0, max_score)
        ax.set_xlabel(f"Notes sur {int(max_score)}", fontsize=10, color=c["muted"])
        ax.set_ylabel("Nombre d\'élèves", fontsize=10, color=c["muted"])
        class_name = self.cmb_hist.currentText()
        ax.set_title(
            f"Distribution des Notes — {class_name}\nAnnée: {year_label}",
            fontsize=12,
            fontweight="bold",
            color=c["text"],
        )
        avg = sum(grades) / len(grades)
        ax.axvline(avg, color=c["danger"], linestyle="--", linewidth=2, label=f"Moyenne: {avg:.1f}/{int(max_score)}")
        ax.legend(facecolor=c["panel"], labelcolor=c["muted"])
        ax.grid(axis="y", alpha=0.25, linestyle="--")
        ax.tick_params(colors=c["muted"])
        for spine in ax.spines.values():
            spine.set_color(c["border"])
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        self.hist_fig.tight_layout()
        self.hist_canvas.draw()

    def _draw_finance_bar(self) -> None:
        c = _chart_colors()
        period = self.cmb_period.currentText()
        self.fin_bar_fig.patch.set_facecolor(c["figure"])
        self.fin_bar_fig.clear()
        ax = self.fin_bar_fig.add_subplot(111)
        ax.set_facecolor(c["panel"])
        try:
            db = DatabaseManager()
            with db.get_connection() as conn:
                repo = AnalyticsRepository(conn)
                income_data = repo.get_monthly_income_totals(period)
                expense_data = repo.get_monthly_expense_totals(period)
        except Exception as exc:
            AppLogger.error("AnalyticsDashboard", f"finance bar: {exc}")
            ax.text(0.5, 0.5, "Erreur de chargement", ha="center", va="center", color=c["danger"], fontsize=12)
            ax.axis("off")
            self.fin_bar_canvas.draw()
            return
        inc_dict = {r[0]: r[1] or 0 for r in income_data}
        exp_dict = {r[0]: r[1] or 0 for r in expense_data}
        months = sorted(set(inc_dict) | set(exp_dict))
        if not months:
            ax.text(0.5, 0.5, "Aucune donnée financière", ha="center", va="center", color=c["muted"], fontsize=12)
            ax.axis("off")
            self.fin_bar_canvas.draw()
            return
        inc = [inc_dict.get(m, 0) for m in months]
        exp = [exp_dict.get(m, 0) for m in months]
        x = range(len(months))
        ww = 0.35
        ax.bar([i - ww / 2 for i in x], inc, ww, label="Recettes", color="#66bb6a")
        ax.bar([i + ww / 2 for i in x], exp, ww, label="Dépenses", color=c["danger"])
        ax.set_xticks(list(x))
        ax.set_xticklabels(months, rotation=45, ha="right")
        ax.set_ylabel("Montant (FCFA)", fontsize=10, color=c["muted"])
        ax.set_title(f"Évolution Financière ({period})", fontsize=11, fontweight="bold", color=c["text"])
        ax.legend(facecolor=c["panel"], labelcolor=c["muted"])
        ax.grid(axis="y", alpha=0.2, linestyle="--")
        ax.tick_params(colors=c["muted"])
        for spine in ax.spines.values():
            spine.set_color(c["border"])
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        self.fin_bar_fig.tight_layout()
        self.fin_bar_canvas.draw()

    def _build_comprehensive_pdf(self, period: str) -> FPDF:
        pdf = FPDF()
        pdf.add_page()
        apply_grades_sheet_header(pdf, get_school_info_row(), "RAPPORT COMPREHENSIF")

        def _lat(text) -> str:
            return ("" if not text else str(text)).encode("latin-1", "ignore").decode("latin-1")

        pdf.set_font("Arial", "", 10)
        pdf.set_text_color(51, 65, 85)
        pdf.cell(0, 7, _lat(f"Période: {period} | Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"), 0, 1, "C")
        pdf.ln(2)
        db = DatabaseManager()
        with db.get_connection() as conn:
            repo = AnalyticsRepository(conn)
            year_id, year_label = repo.get_active_year_context()
            total_inc = repo.get_total_income_by_period(period)
            total_exp = repo.get_total_expense_by_period(period)
            stats = repo.get_comprehensive_stats(year_id)
        balance = total_inc - total_exp
        apply_table_header_style(pdf, "Arial", 10)
        pdf.cell(0, 8, _lat("Résumé Financier"), 1, 1, "L", True)
        apply_table_body_style(pdf, "Arial", 10)
        for idx, (lbl, val) in enumerate(
            [
                ("Total Recettes", f"{total_inc:,.0f} FCFA"),
                ("Total Dépenses", f"{total_exp:,.0f} FCFA"),
                ("Solde Net", f"{balance:,.0f} FCFA"),
            ]
        ):
            set_zebra_row_fill(pdf, idx)
            pdf.cell(110, 8, _lat(lbl), 1, 0, "L", True)
            pdf.cell(80, 8, _lat(val), 1, 1, "R", True)
        pdf.ln(3)
        apply_table_header_style(pdf, "Arial", 10)
        pdf.cell(0, 8, _lat("Résumé Académique"), 1, 1, "L", True)
        apply_table_body_style(pdf, "Arial", 10)
        for idx, (lbl, val) in enumerate(
            [
                ("Année scolaire active", year_label),
                ("Classes", str(stats.get("total_classes", "—"))),
                ("Étudiants actifs", str(stats.get("active_students", "—"))),
                ("Présences", str(stats.get("presents", "—"))),
                ("Absences", str(stats.get("absents", "—"))),
                ("Retards", str(stats.get("lates", "—"))),
            ]
        ):
            set_zebra_row_fill(pdf, idx)
            pdf.cell(110, 8, _lat(lbl), 1, 0, "L", True)
            pdf.cell(80, 8, _lat(val), 1, 1, "R", True)
        return pdf

    def _export_excel(self, report_type: str) -> None:
        names = {
            "financial": "Rapport_Financier",
            "students": "Liste_Etudiants",
            "attendance": "Rapport_Assiduite",
            "grades": "Rapport_Notes",
        }
        default = f"{names.get(report_type, 'Rapport')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Enregistrer le rapport", default, "Excel (*.xlsx)")
        if not path:
            return
        params: dict = {"output_path": path}
        if report_type == "financial":
            params["period"] = self.cmb_period.currentText()
        self.progress_bar.show()
        self.progress_bar.setValue(0)
        self.status_lbl.setText("⏳ Génération en cours…")
        self.status_lbl.setStyleSheet(f"color:{ThemeManager.get_colors().TEXT_SECONDARY}; font-weight:bold;")
        self._report_worker = ReportWorker(f"excel_{report_type}", params)
        self._report_worker.progress.connect(self.progress_bar.setValue)
        self._report_worker.finished.connect(self._on_export_done)
        self._report_worker.error.connect(self._on_export_err)
        self._report_worker.start()
        AppLogger.info("AnalyticsDashboard", f"Export Excel démarré: {report_type}")

    def _export_comprehensive_pdf(self) -> None:
        period = self.cmb_period.currentText()
        try:
            pdf = self._build_comprehensive_pdf(period)
            output_pdf(
                pdf,
                self,
                f"Rapport_Complet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                mode=COMPREHENSIVE_PDF_MODE,
                dialog_title="Enregistrer le rapport PDF",
                success_save_message="Rapport PDF généré avec succès.",
                success_print_message="Rapport envoyé à l\'imprimante.",
            )
            AppLogger.info("AnalyticsDashboard", "PDF compréhensif généré")
        except Exception as exc:
            AppLogger.error("AnalyticsDashboard", f"PDF error: {exc}")
            QMessageBox.critical(self, "Erreur", f"Échec de la création du PDF:\n{exc}")

    def _on_export_done(self, filepath: str) -> None:
        self.progress_bar.hide()
        self.status_lbl.setText(f"✅ Export réussi: {os.path.basename(filepath)}")
        self.status_lbl.setStyleSheet("color:#10b981; font-weight:bold;")
        AppLogger.info("AnalyticsDashboard", f"Export terminé: {filepath}")
        reply = QMessageBox.question(
            self,
            "Export Réussi",
            f"Fichier généré:\n{filepath}\n\nVoulez-vous l\'ouvrir?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            try:
                os.startfile(filepath)
            except Exception as exc:
                AppLogger.warning("AnalyticsDashboard", f"Impossible d\'ouvrir: {exc}")

    def _on_export_err(self, error: str) -> None:
        self.progress_bar.hide()
        self.status_lbl.setText("❌ Erreur lors de l\'export")
        self.status_lbl.setStyleSheet("color:#ef4444; font-weight:bold;")
        QMessageBox.critical(self, "Erreur", f"Erreur lors de l\'export:\n{error}")

    def refresh_data(self) -> None:
        """Called by main_dashbord when this module becomes active."""
        self._load_filters()
        self._refresh()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AnalyticsDashboardWindow()
    window.show()
    sys.exit(app.exec())
